from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles.borders import Border, Side
start_time = datetime.now()

# Help


def octant_analysis(mod=5000):
    from openpyxl import load_workbook
    import os
    import numpy as np
    from operator import itemgetter

    os.system('cls')
    curr = os.getcwd()  
    if os.path.exists("output"):
        for f in os.listdir("output"):  
            os.remove(os.path.join("output", f))
        os.rmdir("output")  # delete output folder
    os.mkdir(curr.replace('\\', '/')+"/output/")  # make output folder
    os.chdir("input")
    fil = os. listdir()  # storing all file name of input folder in list
    os.chdir(curr)
    for list in fil:
        print(list)
        os.chdir("input")  # taking input
        wb = load_workbook(list)
        sheet = wb.active

        ######### making  new lists ##########
        time = []  
        v = []   
        u = []   
        w = []   

        # making  lists for difference
        vd = []    
        ud = []    
        wd = []    
        oct = []    # make a list for octant stori
        wb = load_workbook(list)

        sheet = wb.active
        n = sheet.max_row  
        for r in range(2, n+1):       
            time.append(sheet.cell(row=r, column=1).value)
            u.append(sheet.cell(row=r, column=2).value)
            # making of w list from given data
            v.append(sheet.cell(row=r, column=3).value)
            w.append(sheet.cell(row=r, column=4).value)
        
        ########### Calculating mean v,u,w
        v_mean = np.mean(v, dtype=np.float64)  
        u_mean = np.mean(u, dtype=np.float64) 
        w_mean = np.mean(w, dtype=np.float64) 

        for r in range(2, n+1):
            ud.append(float(sheet.cell(row=r, column=2).value) -
                      u_mean)  
            vd.append(float(sheet.cell(row=r, column=3).value) -
                      v_mean)  
            wd.append(float(sheet.cell(row=r, column=4).value) -
                      w_mean)  
        tu3 = []

        ######### Getting octant     
        for i in range(0, n-1):  
            if ((ud[i] >= 0) and (vd[i] >= 0) and (wd[i] >= 0)):
                oct.append(1)
            elif ((ud[i] >= 0) and (vd[i] >= 0) and (wd[i] < 0)):
                oct.append(-1)
            elif ((ud[i] < 0) and (vd[i] >= 0) and (wd[i] >= 0)):
                oct.append(2)
            elif ((ud[i] < 0) and (vd[i] >= 0) and (wd[i] < 0)):
                oct.append(-2)
            elif ((ud[i] >= 0) and (vd[i] < 0) and (wd[i] >= 0)):
                oct.append(4)
            elif ((ud[i] >= 0) and (vd[i] < 0) and (wd[i] < 0)):
                oct.append(-4)
            elif ((ud[i] < 0) and (vd[i] < 0) and (wd[i] >= 0)):
                oct.append(3)
            else:
                oct.append(-3)
        
        ######## Declaring variables for storing the longest subsequence value of octant -1
        lon_s1 = 0  
        lon_s2 = 0 
        lon_s3 = 0 
        lon_s4 = 0  
        lon_s5 = 0 
        lon_s6 = 0  
        lon_s7 = 0 
        lon_s8 = 0  

        ########### Declaring variables giving the longest subsequence length for octants
        a = 0  
        b = 0  
        c = 0 
        d = 0  
        e = 0 
        f = 0  
        g = 0 
        h = 0  

        ######### Count of octants
        c1 = 0 
        c2 = 0  
        c3 = 0 
        c4 = 0  
        c5 = 0 
        c6 = 0  
        c7 = 0 
        c8 = 0  # It will give the count for octant -4

        # This list will store the starting time for longest subsequence length for octants
 

        for r in range(0, n-1):
            if oct[r] == 1:
                lon_s1 = lon_s1+1
                if r == n-2:
                    if lon_s1 > a:
                        a = lon_s1
                        lon_s1 = 0
                        c1 = 1
                    elif a > lon_s1:
                        lon_s1 = 0
                    else:
                        lon_s1 = 0
                        c1 = c1+1
            else:
                if lon_s1 > a:
                    a = lon_s1
                    lon_s1 = 0
                    c1 = 1
                elif a > lon_s1:
                    lon_s1 = 0
                else:
                    lon_s1 = 0
                    c1 = c1+1
        cn = 0 
		t1 = [] 
        for r in range(0, n-1):
            if oct[r] == 1:
                if cn == 0:
                    ti1 = time[r]
                cn = cn+1
            else:
                cn = 0
            if cn == a:  
                t1.append(ti1)
                cn = 0

                
        for r in range(0, n-1):
            if oct[r] == -1:
                lon_s2 = lon_s2+1
                if r == n-2:
                    if lon_s2 > b:
                        b = lon_s2
                        lon_s2 = 0
                        c2 = 1
                    elif b > lon_s2:
                        lon_s2 = 0
                    else:
                        lon_s2 = 0
                        c2 = c2+1
            else:
                if lon_s2 > b:
                    b = lon_s2
                    lon_s2 = 0
                    c2 = 1
                elif b > lon_s2:
                    lon_s2 = 0
                else:
                    lon_s2 = 0
                    c2 = c2+1

        cn = 0 
		t2 =[] 
        for r in range(0, n-1):
            if oct[r] == -1:
                if cn == 0:
                    ti2 = time[r]
                cn = cn+1
            else:
                cn = 0
            if cn == b:  
                t2.append(ti2)
                cn = 0

        for r in range(0, n-1):
            if oct[r] == 2:
                lon_s3 = lon_s3+1
                if r == n-2:
                    if lon_s3 > c:
                        c = lon_s3
                        lon_s3 = 0
                        c3 = 1
                    elif c > lon_s3:
                        lon_s3 = 0
                    else:
                        lon_s3 = 0
                        c3 = c3+1
            else:
                if lon_s3 > c:
                    c = lon_s3
                    lon_s3 = 0
                    c3 = 1
                elif c > lon_s3:
                    lon_s3 = 0
                else:
                    lon_s3 = 0
                    c3 = c3+1
        cn = 0  
		t3 = []
        for r in range(0, n-1):
            if oct[r] == 2:
                if cn == 0:
                    ti3 = time[r]
                cn = cn+1
            else:
                cn = 0
            if cn == c:  
                t3.append(ti3)
                cn = 0

        for r in range(0, n-1):
            if oct[r] == -2:
                lon_s4 = lon_s4+1
                if r == n-2:
                    if lon_s4 > d:
                        d = lon_s4
                        lon_s4 = 0
                        c4 = 1
                    elif d > lon_s4:
                        lon_s4 = 0
                    else:
                        lon_s4 = 0
                        c4 = c4+1
            else:
                if lon_s4 > d:
                    d = lon_s4
                    lon_s4 = 0
                    c4 = 1
                elif d > lon_s4:
                    lon_s4 = 0
                else:
                    lon_s4 = 0
                    c4 = c4+1
            
        cn = 0 
		t4 = [] 
        for r in range(0, n-1):
            if oct[r] == -2:
                if cn == 0:
                    ti4 = time[r]
                cn = cn+1
            else:
                cn = 0
            if cn == d:  
                t4.append(ti4)
                cn = 0

        for r in range(0, n-1):
            if oct[r] == 3:
                lon_s5 = lon_s5+1
                if r == n-2:
                    if lon_s5 > e:
                        e = lon_s5
                        lon_s5 = 0
                        c5 = 1
                    elif e > lon_s5:
                        lon_s5 = 0
                    else:
                        lon_s5 = 0
                        c5 = c5+1
            else:
                if lon_s5 > e:
                    e = lon_s5
                    lon_s5 = 0
                    c5 = 1
                elif e > lon_s5:
                    lon_s5 = 0
                else:
                    lon_s5 = 0
                    c5 = c5+1
            
        cn = 0  
		t5 = []
        for r in range(0, n-1):
            if oct[r] == 3:
                if cn == 0:
                    ti5 = time[r]
                cn = cn+1
            else:
                cn = 0
            if cn == e:  
                t5.append(ti5)
                cn = 0

        for r in range(0, n-1):
            if oct[r] == -3:
                lon_s6 = lon_s6+1
                if r == n-2:
                    if lon_s6 > a:
                        f = lon_s6
                        lon_s6 = 0
                        c6 = 1
                    elif f > lon_s6:
                        lon_s6 = 0
                    else:
                        lon_s6 = 0
                        c6 = c6+1
            else:
                if lon_s6 > f:
                    f = lon_s6
                    lon_s6 = 0
                    c6 = 1
                elif f > lon_s6:
                    lon_s6 = 0
                else:
                    lon_s6 = 0
                    c6 = c6+1
            
        cn = 0  
		t6 = []
        for r in range(0, n-1):
            if oct[r] == -3:
                if cn == 0:
                    ti6 = time[r]
                cn = cn+1
            else:
                cn = 0
            if cn == f:  
                t6.append(ti6)
                cn = 0

        for r in range(0, n-1):
            if oct[r] == 4:
                lon_s7 = lon_s7+1
                if r == n-2:
                    if lon_s7 > g:
                        g = lon_s7
                        lon_s7 = 0
                        c7 = 1
                    elif g > lon_s7:
                        lon_s7 = 0
                    else:
                        lon_s7 = 0
                        c7 = c7+1
            else:
                if lon_s7 > g:
                    g = lon_s7
                    lon_s7 = 0
                    c7 = 1
                elif g > lon_s7:
                    lon_s7 = 0
                else:
                    lon_s7 = 0
                    c7 = c7+1
        cn = 0  
		t7 = []
        for r in range(0, n-1):
            if oct[r] == 4:
                if cn == 0:
                    ti7 = time[r]
                cn = cn+1
            else:
                cn = 0
            if cn == g:  
                t7.append(ti7)
                cn = 0

        for r in range(0, n-1):
            if oct[r] == -4:
                lon_s8 = lon_s8+1
                if r == n-2:
                    if lon_s8 > a:
                        h = lon_s8
                        lon_s8 = 0
                        c8 = 1
                    elif h > lon_s8:
                        lon_s8 = 0
                    else:
                        lon_s8 = 0
                        c8 = c8+1
            else:
                if lon_s8 > h:
                    h = lon_s8
                    lon_s8 = 0
                    c8 = 1
                elif h > lon_s8:
                    lon_s8 = 0
                else:
                    lon_s8 = 0
                    c8 = c8+1
        cn = 0 
		t8 = [] 
        for r in range(0, n-1):
            if oct[r] == -4:
                if cn == 0:
                    ti8 = time[r]
                cn = cn+1
            else:
                cn = 0
            if cn == h:  
                t8.append(ti8)
                cn = 0

        tu3.append(["Octant##", "Longest Subsequence length", "Count"])
        tu3.append(["1", a, c1])
        tu3.append(["-1", b, c2])
        tu3.append(["2", c, c3])
        tu3.append(["-2", d, c4])
        tu3.append(["3", e, c5])
        tu3.append(["-3", f, c6])
        tu3.append(["4", g, c7])
        tu3.append(["-4", h, c8])
        for i in range(0, 1000):
            tu3.append(["", "", ""])
        
        ########  printing the first column the new requireduired table
        col_1 = []  
        col_1.append("Octant###")
        col_1.append("1")
        col_1.append("Time")
        for i in range(0, c1):  
            col_1.append("")
        col_1.append("-1")
        col_1.append("Time")
        for i in range(0, c2):  
            col_1.append("")
        col_1.append("2")
        col_1.append("Time")
        for i in range(0, c3):  
            col_1.append("")
        col_1.append("-2")
        col_1.append("Time")
        for i in range(0, c4):  
            col_1.append("")
        col_1.append("3")
        col_1.append("Time")
        for i in range(0, c5):  
            col_1.append("")
        col_1.append("-3")
        col_1.append("Time")
        for i in range(0, c6):  
            col_1.append("")
        col_1.append("4")
        col_1.append("Time")
        for i in range(0, c7):  
            col_1.append("")
        col_1.append("-4")
        col_1.append("Time")
        for i in range(0, c8):  
            col_1.append("")
        v = len(col_1)

        ########printing the second coloumn new requireduired table
        col_2 = []  
        col_2.append("Longest Subsequence Length")
        col_2.append(a)
        col_2.append("From")
        
        for i in range(0, c1):
            col_2.append(t1[i])
        col_2.append(b)
        col_2.append("From")
        
        for i in range(0, c2):
            col_2.append(t2[i])
        col_2.append(c)
        col_2.append("From")
        
        for i in range(0, c3):
            col_2.append(t3[i])
        col_2.append(d)
        col_2.append("From")
        
        for i in range(0, c4):
            col_2.append(t4[i])
        col_2.append(e)
        col_2.append("From")
        
        for i in range(0, c5):
            col_2.append(t5[i])
        col_2.append(f)
        col_2.append("From")
        
        for i in range(0, c6):
            col_2.append(t6[i])
        col_2.append(g)
        col_2.append("From")

        for i in range(0, c7):
            col_2.append(t7[i])
        col_2.append(h)
        col_2.append("From")

        for i in range(0, c8):
            col_2.append(t8[i])

        ########  printing third coloumn the new requireduired table
        col_3 = []  
        col_3.append("Count")
        col_3.append(c1)
        col_3.append("To")
        
        for i in range(0, c1):
            col_3.append(t1[i] + 0.01*(a-1))
        col_3.append(c2)
        col_3.append("To")
        
        for i in range(0, c2):
            col_3.append(t2[i] + 0.01*(b-1))
        col_3.append(c3)
        col_3.append("To")
        
        for i in range(0, c3):
            col_3.append(t3[i] + 0.01*(c-1))
        col_3.append(c4)
        col_3.append("To")
        
        for i in range(0, c4):
            col_3.append(t4[i] + 0.01*(d-1))
        col_3.append(c5)
        col_3.append("To")
        
        for i in range(0, c5):
            col_3.append(t5[i] + 0.01*(e-1))
        col_3.append(c6)
        col_3.append("To")
        
        for i in range(0, c6):
            col_3.append(t6[i] + 0.01*(f-1))
        col_3.append(c7)
        col_3.append("To")
        
        for i in range(0, c7):
            col_3.append(t7[i] + 0.01*(g-1))
        col_3.append(c8)
        col_3.append("To")
        
        for i in range(0, c8):
            col_3.append(t8[i] + 0.01*(h-1))

        for i in range(0, 1000):
            col_1.append("")
            col_2.append("")
            col_3.append("")

        ##### for overall count######### 
        c1 = 0  
        c2 = 0  
        c3 = 0 
        c4 = 0  
        c5 = 0 
        c6 = 0  
        c7 = 0 
        c8 = 0  

        ################# Getting octants ###########
        for i in range(0, n-1):  
            if (oct[i] == 1):
                c1 = c1+1
            elif (oct[i] == -1):
                c2 = c2+1
            elif (oct[i] == 2):
                c3 = c3+1
            elif (oct[i] == -2):
                c4 = c4+1
            elif (oct[i] == 3):
                c5 = c5+1
            elif (oct[i] == -3):
                c6 = c6+1
            elif (oct[i] == 4):
                c7 = c7+1
            else:
                c8 = c8+1
        rank = []  # this list will give the ranks of all octants in differnt intervals
        rank1_id_name = []  # This list will help rank list in storing the data
        required = []  # This list will help rank list in storing the data
        required.append([c1, 1])
        required.append([c2, 2])
        required.append([c3, 3])
        required.append([c4, 4])
        required.append([c5, 5])
        required.append([c6, 6])
        required.append([c7, 7])
        required.append([c8, 8])
        required.sort()


        ########### Giving ranks in sorted list
        required_ranks = [0, 0, 0, 0, 0, 0, 0, 0]
        required_ranks[required[0][1]-1] = 8
        required_ranks[required[1][1]-1] = 7
        required_ranks[required[2][1]-1] = 6
        required_ranks[required[3][1]-1] = 5
        required_ranks[required[4][1]-1] = 4
        required_ranks[required[5][1]-1] = 3
        required_ranks[required[6][1]-1] = 2
        required_ranks[required[7][1]-1] = 1
        rank.append(required_ranks)  # Appending this list in rank list
        rank_id_name = []

        ############## # Givng name and id to Rank 1 Octant
        if required_ranks[0] == 1: 
            rank_id_name = [1, "Internal outward Interaction"]
            rank1_id_name.append(rank_id_name)
        elif required_ranks[1] == 1:  
            rank_id_name = [-1, "External outward Interaction"]
            rank1_id_name.append(rank_id_name)
        elif required_ranks[2] == 1:  
            rank_id_name = [2, "External Ejection"]
            rank1_id_name.append(rank_id_name)
        elif required_ranks[3] == 1:  
            rank_id_name = [-2, "Internal Ejection"]
            rank1_id_name.append(rank_id_name)
        elif required_ranks[4] == 1:  
            rank_id_name = [3, "External inward Interaction"]
            rank1_id_name.append(rank_id_name)
        elif required_ranks[5] == 1:  
            rank_id_name = [-3, "Internal inward Interaction"]
            rank1_id_name.append(rank_id_name)
        elif required_ranks[6] == 1:  
            rank_id_name = [4, "Internal Sweep"]
            rank1_id_name.append(rank_id_name)
        elif required_ranks[7] == 1:  
            rank_id_name = [-4, "External Sweep"]
            rank1_id_name.append(rank_id_name)

        ############ Declaring list of octant count in mod range 
        oc1 = []  
        oc2 = []  
        oc3 = []  
        oc4 = []  
        oc5 = []  
        oc6 = []  
        oc7 = []  
        oc8 = []  
        ########## Declaring variables for using in function
        oct1 = 0  
        oct2 = 0  
        oct3 = 0  
        oct4 = 0  
        oct5 = 0  
        oct6 = 0  
        oct7 = 0  
        oct8 = 0  
        m = int((n-2)/mod) + 1  # m will give no.of interval

        ###########Declaring variables for Count of rank 1 if Octants in different mod intervals
        orc1 = 0  
        orc2 = 0  
        orc3 = 0
        orc4 = 0  
        orc5 = 0 
        orc6 = 0  
        orc7 = 0 
        orc8 = 0
        
        ######## calculating no. of octants in particular a interval
        for i in range(m):
            s = mod*i
            for j in range(mod):
                if (j+s < n-1):
                    if (oct[j+s] == 1):
                        oct1 = oct1+1  
                    elif (oct[j+s] == -1):
                        oct2 = oct2+1  
                    elif (oct[j+s] == 2):
                        oct3 = oct3+1 
                    elif (oct[j+s] == -2):
                        oct4 = oct4+1  
                    elif (oct[j+s] == 3):
                        oct5 = oct5+1 
                    elif (oct[j+s] == -3):
                        oct6 = oct6+1   
                    elif (oct[j+s] == 4):
                        oct7 = oct7+1 
                    elif (oct[j+s] == -4):
                        oct8 = oct8+1  

            oc1.append(oct1)
            oc2.append(oct2)
            oc3.append(oct3)
            oc4.append(oct4)
            oc5.append(oct5)
            oc6.append(oct6)
            oc7.append(oct7)
            oc8.append(oct8)
            required = []  # This list is used for storing the rank in a particular interval
            required.append([oct1, 1])
            required.append([oct2, 2])
            required.append([oct3, 3])
            required.append([oct4, 4])
            required.append([oct5, 5])
            required.append([oct6, 6])
            required.append([oct7, 7])
            required.append([oct8, 8])
            required.sort()  # Sorting required with respect to First Coloumn
            
            ###### Giving ranks in sorted list
            required_ranks = [0, 0, 0, 0, 0, 0, 0, 0]
            required_ranks[required[0][1]-1] = 8  
            required_ranks[required[1][1]-1] = 7  
            required_ranks[required[2][1]-1] = 6  
            required_ranks[required[3][1]-1] = 5  
            required_ranks[required[4][1]-1] = 4  
            required_ranks[required[5][1]-1] = 3  
            required_ranks[required[6][1]-1] = 2  
            required_ranks[required[7][1]-1] = 1  
            rank.append(required_ranks)
            rank_id_name = []
            rank_id_name = []

            ############# Givng name and id to Rank 1 Octant
            if required_ranks[0] == 1:  
                rank_id_name = [1, "Internal outward Interaction"]
                rank1_id_name.append(rank_id_name)
                orc1 = orc1+1
            elif required_ranks[1] == 1:  
                rank_id_name = [-1, "External outward Interaction"]
                rank1_id_name.append(rank_id_name)
                orc2 = orc2+1
            elif required_ranks[2] == 1:  
                rank_id_name = [2, "External Ejection"]
                rank1_id_name.append(rank_id_name)
                orc3 = orc3+1
            elif required_ranks[3] == 1:  
                rank_id_name = [-2, "Internal Ejection"]
                rank1_id_name.append(rank_id_name)
                orc4 = orc4+1
            elif required_ranks[4] == 1:  
                rank_id_name = [3, "External inward Interaction"]
                rank1_id_name.append(rank_id_name)
                orc5 = orc5+1
            elif required_ranks[5] == 1:  
                rank_id_name = [-3, "Internal inward Interaction"]
                rank1_id_name.append(rank_id_name)
                orc6 = orc6+1
            elif required_ranks[6] == 1:  
                rank_id_name = [4, "Internal Sweep"]
                rank1_id_name.append(rank_id_name)
                orc7 = orc7+1
            elif required_ranks[7] == 1:  
                rank_id_name = [-4, "External Sweep"]
                rank1_id_name.append(rank_id_name)
                orc8 = orc8+1
            oct1 = 0
            oct2 = 0
            oct3 = 0
            oct4 = 0
            oct5 = 0
            oct6 = 0
            oct7 = 0
            oct8 = 0

        r = []  # it is giving full data that we need to print
        re = []  # it will give the data 2-d matrix of respective interval
        over = [[0]*8, [0]*8, [0]*8, [0]*8, [0]*8, [0]*8, [0]
                * 8, [0]*8]  # This is matirx of overall transition
        for i in range(0, m):
            p = i*mod

            ######### Declaring list for storing the value each element in 2-d matrix of different intervals
            a1 = [0]*8
            a2 = [0]*8
            a3 = [0]*8
            a4 = [0]*8
            a5 = [0]*8
            a6 = [0]*8
            a7 = [0]*8
            a8 = [0]*8

            for j in range(0, mod):
                if j+p <= n-3:
                    if oct[j+p] == 1 and oct[j+p+1] == 1:
                        a1[0] = a1[0]+1
                        # This is for overall transitin matrix
                        over[0][0] = over[0][0]+1
                    elif oct[j+p] == 1 and oct[j+p+1] == -1:
                        # This will keep on increasing the value of each transition in different intervals
                        a1[1] = a1[1]+1
                        # This is for overall transitin matrix
                        over[0][1] = over[0][1]+1
                    elif oct[j+p] == 1 and oct[j+p+1] == 2:
                        a1[2] = a1[2]+1
                        over[0][2] = over[0][2]+1
                    elif oct[j+p] == 1 and oct[j+p+1] == -2:
                        a1[3] = a1[3]+1
                        over[0][3] = over[0][3]+1
                    elif oct[j+p] == 1 and oct[j+1+p] == 3:
                        a1[4] = a1[4]+1
                        over[0][4] = over[0][4]+1
                    elif oct[j+p] == 1 and oct[j+p+1] == -3:
                        a1[5] = a1[5]+1
                        over[0][5] = over[0][5]+1
                    elif oct[j+p] == 1 and oct[j+1+p] == 4:
                        a1[6] = a1[6]+1
                        over[0][6] = over[0][6]+1
                    elif oct[j+p] == 1 and oct[j+p+1] == -4:
                        a1[7] = a1[7]+1
                        over[0][7] = over[0][7]+1
                    elif oct[j+p] == -1 and oct[j+p+1] == 1:
                        a2[0] = a2[0]+1
                        over[1][0] = over[1][0]+1
                    elif oct[j+p] == -1 and oct[j+p+1] == -1:
                        a2[1] = a2[1]+1
                        over[1][1] = over[1][1]+1
                    elif oct[j+p] == -1 and oct[j+p+1] == 2:
                        a2[2] = a2[2]+1
                        over[1][2] = over[1][2]+1
                    elif oct[j+p] == -1 and oct[j+p+1] == -2:
                        a2[3] = a2[3]+1
                        over[1][3] = over[1][3]+1
                    elif oct[j+p] == -1 and oct[j+p+1] == 3:
                        a2[4] = a2[4]+1
                        over[1][4] = over[1][4]+1
                    elif oct[j+p] == -1 and oct[j+1+p] == -3:
                        a2[5] = a2[5]+1
                        over[1][5] = over[1][5]+1
                    elif oct[j+p] == -1 and oct[j+p+1] == 4:
                        a2[6] = a2[6]+1
                        over[1][6] = over[1][6]+1
                    elif oct[j+p] == -1 and oct[j+1+p] == -4:
                        a2[7] = a2[7]+1
                        over[1][7] = over[1][7]+1
                    elif oct[j+p] == 2 and oct[j+p+1] == 1:
                        a3[0] = a3[0]+1
                        over[2][0] = over[2][0]+1
                    elif oct[j+p] == 2 and oct[j+p+1] == -1:
                        a3[1] = a3[1]+1
                        over[2][1] = over[2][1]+1
                    elif oct[j+p] == 2 and oct[j+p+1] == 2:
                        a3[2] = a3[2]+1
                        over[2][2] = over[2][2]+1
                    elif oct[j+p] == 2 and oct[j+p+1] == -2:
                        a3[3] = a3[3]+1
                        over[2][3] = over[2][3]+1
                    elif oct[j+p] == 2 and oct[j+p+1] == 3:
                        a3[4] = a3[4]+1
                        over[2][4] = over[2][4]+1
                    elif oct[j+p] == 2 and oct[j+1+p] == -3:
                        a3[5] = a3[5]+1
                        over[2][5] = over[2][5]+1
                    elif oct[j+p] == 2 and oct[j+p+1] == 4:
                        a3[6] = a3[6]+1
                        over[2][6] = over[2][6]+1
                    elif oct[j+p] == 2 and oct[j+1+p] == -4:
                        a3[7] = a3[7]+1
                        over[2][7] = over[2][7]+1
                    elif oct[j+p] == -2 and oct[j+p+1] == 1:
                        a4[0] = a4[0]+1
                        over[3][0] = over[3][0]+1
                    elif oct[j+p] == -2 and oct[j+p+1] == -1:
                        a4[1] = a4[1]+1
                        over[3][1] = over[3][1]+1
                    elif oct[j+p] == -2 and oct[j+p+1] == 2:
                        a4[2] = a4[2]+1
                        over[3][2] = over[3][2]+1
                    elif oct[j+p] == -2 and oct[j+p+1] == -2:
                        a4[3] = a4[3]+1
                        over[3][3] = over[3][3]+1
                    elif oct[j+p] == -2 and oct[j+p+1] == 3:
                        a4[4] = a4[4]+1
                        over[3][4] = over[3][4]+1
                    elif oct[j+p] == -2 and oct[j+1+p] == -3:
                        a4[5] = a4[5]+1
                        over[3][5] = over[3][5]+1
                    elif oct[j+p] == -2 and oct[j+p+1] == 4:
                        a4[6] = a4[6]+1
                        over[3][6] = over[3][6]+1
                    elif oct[j+p] == -2 and oct[j+1+p] == -4:
                        a4[7] = a4[7]+1
                        over[3][7] = over[3][7]+1
                    elif oct[j+p] == 3 and oct[j+p+1] == 1:
                        a5[0] = a5[0]+1
                        over[4][0] = over[4][0]+1
                    elif oct[j+p] == 3 and oct[j+p+1] == -1:
                        a5[1] = a5[1]+1
                        over[4][1] = over[4][1]+1
                    elif oct[j+p] == 3 and oct[j+p+1] == 2:
                        a5[2] = a5[2]+1
                        over[4][2] = over[4][2]+1
                    elif oct[j+p] == 3 and oct[j+p+1] == -2:
                        a5[3] = a5[3]+1
                        over[4][3] = over[4][3]+1
                    elif oct[j+p] == 3 and oct[j+p+1] == 3:
                        a5[4] = a5[4]+1
                        over[4][4] = over[4][4]+1
                    elif oct[j+p] == 3 and oct[j+1+p] == -3:
                        a5[5] = a5[5]+1
                        over[4][5] = over[4][5]+1
                    elif oct[j+p] == 3 and oct[j+p+1] == 4:
                        a5[6] = a5[6]+1
                        over[4][6] = over[4][6]+1
                    elif oct[j+p] == 3 and oct[j+1+p] == -4:
                        a5[7] = a5[7]+1
                        over[4][7] = over[4][7]+1
                    elif oct[j+p] == -3 and oct[j+p+1] == 1:
                        a6[0] = a6[0]+1
                        over[5][0] = over[5][0]+1
                    elif oct[j+p] == -3 and oct[j+p+1] == -1:
                        a6[1] = a6[1]+1
                        over[5][1] = over[5][1]+1
                    elif oct[j+p] == -3 and oct[j+p+1] == 2:
                        a6[2] = a6[2]+1
                        over[5][2] = over[5][2]+1
                    elif oct[j+p] == -3 and oct[j+p+1] == -2:
                        a6[3] = a6[3]+1
                        over[5][3] = over[5][3]+1
                    elif oct[j+p] == -3 and oct[j+p+1] == 3:
                        a6[4] = a6[4]+1
                        over[5][4] = over[5][4]+1
                    elif oct[j+p] == -3 and oct[j+1+p] == -3:
                        a6[5] = a6[5]+1
                        over[5][5] = over[5][5]+1
                    elif oct[j+p] == -3 and oct[j+p+1] == 4:
                        a6[6] = a6[6]+1
                        over[5][6] = over[5][6]+1
                    elif oct[j+p] == -3 and oct[j+1+p] == -4:
                        a6[7] = a6[7]+1
                        over[5][7] = over[5][7]+1
                    elif oct[j+p] == 4 and oct[j+p+1] == 1:
                        a7[0] = a7[0]+1
                        over[6][0] = over[6][0]+1
                    elif oct[j+p] == 4 and oct[j+p+1] == -1:
                        a7[1] = a7[1]+1
                        over[6][1] = over[6][1]+1
                    elif oct[j+p] == 4 and oct[j+p+1] == 2:
                        a7[2] = a7[2]+1
                        over[6][2] = over[6][2]+1
                    elif oct[j+p] == 4 and oct[j+p+1] == -2:
                        a7[3] = a7[3]+1
                        over[6][3] = over[6][3]+1
                    elif oct[j+p] == 4 and oct[j+p+1] == 3:
                        a7[4] = a7[4]+1
                        over[6][4] = over[6][4]+1
                    elif oct[j+p] == 4 and oct[j+1+p] == -3:
                        a7[5] = a7[5]+1
                        over[6][5] = over[6][5]+1
                    elif oct[j+p] == 4 and oct[j+p+1] == 4:
                        a7[6] = a7[6]+1
                        over[6][6] = over[6][6]+1
                    elif oct[j+p] == 4 and oct[j+1+p] == -4:
                        a7[7] = a7[7]+1
                        over[6][7] = over[6][7]+1
                    elif oct[j+p] == -4 and oct[j+p+1] == 1:
                        a8[0] = a8[0]+1
                        over[7][0] = over[7][0]+1
                    elif oct[j+p] == -4 and oct[j+p+1] == -1:
                        a8[1] = a8[1]+1
                        over[7][1] = over[7][1]+1
                    elif oct[j+p] == -4 and oct[j+p+1] == 2:
                        a8[2] = a8[2]+1
                        over[7][2] = over[7][2]+1
                    elif oct[j+p] == -4 and oct[j+p+1] == -2:
                        a8[3] = a8[3]+1
                        over[7][3] = over[7][3]+1
                    elif oct[j+p] == -4 and oct[j+p+1] == 3:
                        a8[4] = a8[4]+1
                        over[7][4] = over[7][4]+1
                    elif oct[j+p] == -4 and oct[j+1+p] == -3:
                        a8[5] = a8[5]+1
                        over[7][5] = over[7][5]+1
                    elif oct[j+p] == -4 and oct[j+p+1] == 4:
                        a8[6] = a8[6]+1
                        over[7][6] = over[7][6]+1
                    elif oct[j+p] == -4 and oct[j+1+p] == -4:
                        a8[7] = a8[7]+1
                        over[7][7] = over[7][7]+1
            if i == 0:
                r.append(over)

            # Making  a 2-d Matrix required by appending eight 1-d matrix
            re.append(a1)
            re.append(a2)
            re.append(a3)
            re.append(a4)
            re.append(a5)
            re.append(a6)
            re.append(a7)
            re.append(a8)
            c = re.copy()
            r.append(c)  # appending the required matrix in r
            re.clear()  # clearing required matrix so that new data may get store into it
        h = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
        row = []
        j = 0
        a = 0
        b = 0
        for x in range(n-2):
            if x == 0:
                continue
            if (x == 1):
                s = "mod "+str(mod)
            elif (x >= 2 and x < 2+m):
                if (x == 1+m):  # it will work only if x=1+m
                    z = j*mod
                    y = (j+1)*mod
                    s = str(z)+"-"+str(n-2)
                    j = j+1
                else:
                    z = j*mod
                    y = (j+1)*mod-1
                    s = str(z)
                    j = j+1
            elif (x-(2+m)) % 9 == 0 and x < 9*(m+1)+2+m:
                if x == 2+m:
                    row.append(["", "Overall transition Count",
                               "", "", "", "", "", "", "", ""])
                    row.append(["", "To", "", "", "", "", "", "", "", ""])
                    row.append(["", "Octant#", "+1", "-1", "+2",
                               "-2", "+3", "-3", "+4", "-4"])
                else:
                    row.append(["", "", "", "", "", "", "", "", "", ""])
                    row.append(["", "", "", "", "", "", "", "", "", ""])
                    row.append(["", "Mod transition Count",
                               "", "", "", "", "", "", "", ""])
                    row.append(["", str(((x-(2+m))//9-1)*mod)+"-"+str(np.minimum(
                        ((x-(2+m))//9)*mod-1, n-2)), "To", "", "", "", "", "", "", ""])
                    row.append(["", "Octant#", "+1", "-1", "+2",
                               "-2", "+3", "-3", "+4", "-4"])
            elif x < 9*(m+1)+(2+m):
                if (x-(2+m)) % 9 == 1:
                    row.append(["From", h[((x-(2+m)) % 9)-1], r[b][a][0], r[b][a][1], r[b]
                               [a][2], r[b][a][3], r[b][a][4], r[b][a][5], r[b][a][6], r[b][a][7]])
                else:
                    row.append(["", h[((x-(2+m)) % 9)-1], r[b][a][0], r[b][a][1], r[b][a]
                               [2], r[b][a][3], r[b][a][4], r[b][a][5], r[b][a][6], r[b][a][7]])
                a = a+1
                if a == 8:
                    a = 0
                    b = b+1
            else:
                row.append(["", "", "", "", "", "", "", "", "", ""])
        z = len(row)