

from datetime import datetime
start_time = datetime.now()

#Help https://youtu.be/N6PBd4XdnEw
def octant_range_names(mod=5000):
 from openpyxl import load_workbook
 import os
 import numpy as np
 os.system('cls')
try:

 time = []  # new list for time
 v = []     # new list for V
 u = []     # new list for U
 w = []     # new list for W
 vd = []    # list for difference of Vavg - V
 ud = []    # list for difference of Uavg -U
 wd = []    # list for difference of Wavg - W
 oct = []   # list for storing octant

 lwb=load_workbook("octant_input.xlsx")
 sheet=lwb["Sheet1"] 
 n=sheet.max_row
 for r in range(2,n+1):
  time.append(sheet.cell(row=r,column=1).value)
  u.append(sheet.cell(row=r,column=2).value)
  v.append(sheet.cell(row=r,column=3).value)
  w.append(sheet.cell(row=r,column=4).value)

 v_avg=np.mean(v, dtype=np.float64) #  calculate average of v
 u_avg=np.mean(u, dtype=np.float64) #  calculate average of u
 w_avg=np.mean(w, dtype=np.float64) #  calculate average of w

 
 for r in range(2,n+1):
  ud.append(float(sheet.cell(row=r,column=2).value)-u_avg)  # push row[1] - u_avg in ud list
  vd.append(float(sheet.cell(row=r,column=3).value)-v_avg)  # push row[2] - v_avg in vd list
  wd.append(float(sheet.cell(row=r,column=4).value)-w_avg)	 # push row[3] - w_avg in wd list 


 for i in range(0,n-1):# find octants
  if ((ud[i]>=0) and (vd[i]>=0) and (wd[i]>=0 )):
   oct.append(1)      # adding the octant in oct list after checking the condition of octant 1 to calculate overall octant 1 value
  elif((ud[i]>=0) and (vd[i]>=0) and (wd[i]<0 )):
   oct.append(-1)     # adding the octant in oct list after checking the condition of octant -1 to calculate overall octant -1 value
  elif((ud[i]<0) and (vd[i]>=0) and (wd[i]>=0 )):
   oct.append(2)      # adding the octant in oct list after checking the condition of octant 2 to calculate overall octant 2 value
  elif((ud[i]<0) and (vd[i]>=0) and (wd[i]<0 )):
   oct.append(-2)     # adding the octant in oct list after checking the condition of octant -2 to calculate overall octant -2 value
  elif((ud[i]>=0) and (vd[i]<0) and (wd[i]>=0 )):
   oct.append(4)      # adding the octant in oct list after checking the condition of octant 4 to calculate overall octant 4 value
  elif((ud[i]>=0) and (vd[i]<0) and (wd[i]<0 )):
   oct.append(-4)    # adding the octant in oct list after checking the condition of octant -4 to calculate overall octant -4 value
  elif((ud[i]<0) and (vd[i]<0) and (wd[i]>=0 )):
   oct.append(3)     # adding the octant in oct list after checking the condition of octant 3 to calculate overall octant 3 value
  else:
   oct.append(-3)    # adding the octant in oct list after checking the condition of octant -3 to calculate overall octant =3 value
     
 count1=0         # calculating total count of octant 1
 count3=0         # calculating total count of octant 2
 count2=0         # calculating total count of octant -1
 count4=0         # calculating total count of octant -2
 count5=0         # calculating total count of octant 3
 count6=0         # calculating total count of octant -3
 count7=0         # calculating total count of octant 4
 count8=0         # calculating total count of octant -4	 
 
 for i in range(0,n-1):# for loop to calculate count of octants
  if (oct[i]==1):
   count1=count1+1
  elif(oct[i]==-1):
   count2=count2+1
  elif(oct[i]==2):
   count3=count3+1
  elif(oct[i]==-2):
   count4=count4+1
  elif(oct[i]==3):
   count5=count5+1
  elif(oct[i]==-3):
   count6=count6+1    
  elif(oct[i]==4):
   count7=count7+1
  else:
   count8=count8+1

 octant_rank=[]    # list to store rank of overall octant count
 octant_rank1_id=[] # list to store rank of overall octant 1 count
 req_count=[]
 req_count.append([count1,1])
 req_count.append([count2,2])
 req_count.append([count3,3])
 req_count.append([count4,4])
 req_count.append([count5,5])
 req_count.append([count6,6])
 req_count.append([count7,7])
 req_count.append([count8,8]) 
 req_count.sort()
 req_octant_ranks=[0,0,0,0,0,0,0,0]
 req_octant_ranks[req_count[0][1]-1]=8
 req_octant_ranks[req_count[1][1]-1]=7
 req_octant_ranks[req_count[2][1]-1]=6
 req_octant_ranks[req_count[3][1]-1]=5
 req_octant_ranks[req_count[4][1]-1]=4
 req_octant_ranks[req_count[5][1]-1]=3
 req_octant_ranks[req_count[6][1]-1]=2
 req_octant_ranks[req_count[7][1]-1]=1
 octant_rank.append(req_octant_ranks)
 octant_rank_id_name=[]

 if req_octant_ranks[0]==1:
  octant_rank_id_name=[1,"Internal outward Interaction"]
  octant_rank1_id.append(octant_rank_id_name)
 elif req_octant_ranks[1]==1:
  octant_rank_id_name=[-1,"External outward Interaction"]
  octant_rank1_id.append(octant_rank_id_name)
 elif req_octant_ranks[2]==1:
  octant_rank_id_name=[2,"External Ejection"]
  octant_rank1_id.append(octant_rank_id_name) 
 elif req_octant_ranks[3]==1:
  octant_rank_id_name=[-2,"Internal Ejection"]
  octant_rank1_id.append(octant_rank_id_name) 
 elif req_octant_ranks[4]==1:
  octant_rank_id_name=[3,"External inward Interaction"]
  octant_rank1_id.append(octant_rank_id_name) 
 elif req_octant_ranks[5]==1:
  octant_rank_id_name=[-3,"Internal inward Interaction"]
  octant_rank1_id.append(octant_rank_id_name) 
 elif req_octant_ranks[6]==1:
  octant_rank_id_name=[4,"Internal Sweep"]
  octant_rank1_id.append(octant_rank_id_name) 
 elif req_octant_ranks[7]==1:
  octant_rank_id_name=[-4,"External Sweep"]
  octant_rank1_id.append(octant_rank_id_name) 

 octant1=[] #list of octant 1
 octant2=[] #list of octant -1
 octant3=[] #list of octant 2
 octant4=[] #list of octant -2
 octant5=[] #list of octant 3
 octant6=[] #list of octant -3
 octant7=[] #list of octant 4
 octant8=[] #list of octant -4

 oct1=0 #making variables for using in function 
 oct2=0 #making variables for using in function
 oct3=0 #making variables for using in function
 oct4=0 #making variables for using in function
 oct5=0 #making variables for using in function
 oct6=0 #making variables for using in function
 oct7=0 #making variables for using in function
 oct8=0 #making variables for using in function

 m= int((n-2)/mod) +1

 r1c=0    #count of rank 1 mod values of octant 1 
 r2c=0    #count of rank 1 mod values of octant -1
 r3c=0    #count of rank 1 mod values of octant 2
 r4c=0    #count of rank 1 mod values of octant -2
 r5c=0    #count of rank 1 mod values of octant 3
 r6c=0    #count of rank 1 mod values of octant -3
 r7c=0    #count of rank 1 mod values of octant 4
 r8c=0    #count of rank 1 mod values of octant -4

 for i in range(m):   # for loop to calculate individual count of octants in range of mod
  s=mod*i
  for j in range(mod):	
   if(j+s<n-1):	                                    
    if (oct[j+s]==1):
     oct1=oct1+1       #calculate count of octant 1 in range of particular mod
    elif(oct[j+s]==-1):
     oct2=oct2+1       #calculate count of octant -1 in range of particular mod
    elif(oct[j+s]==2):
     oct3=oct3+1       #calculate count of octant 2 in range of particular mod
    elif(oct[j+s]==-2):
     oct4=oct4+1       #calculate count of octant -2 in range of particular mod
    elif(oct[j+s]==3):
     oct5=oct5+1       #calculate count of octant 3 in range of particular mod
    elif(oct[j+s]==-3):
     oct6=oct6+1       #calculate count of octant -3 in range of particular mod
    elif(oct[j+s]==4):
     oct7=oct7+1       #calculate count of octant 4 in range of particular mod
    elif(oct[j+s]==-4):
     oct8=oct8+1       #calculate count of octant -4 in range of particular mod
  octant1.append(oct1)  # push values of count of octant 1 in particular range of mod
  octant2.append(oct2)  # push values of count of octant -1 in particular range of mod
  octant3.append(oct3)  # push values of count of octant 2 in particular range of mod
  octant4.append(oct4)  # push values of count of octant -2 in particular range of mod
  octant5.append(oct5)  # push values of count of octant 3 in particular range of mod
  octant6.append(oct6)  # push values of count of octant -3 in particular range of mod
  octant7.append(oct7)  # push values of count of octant 4 in particular range of mod
  octant8.append(oct8)  # push values of count of octant -4 in particular range of mod

  req_count=[]
  req_count.append([oct1,1])
  req_count.append([oct2,2])
  req_count.append([oct3,3])
  req_count.append([oct4,4])
  req_count.append([oct5,5])
  req_count.append([oct6,6])
  req_count.append([oct7,7])
  req_count.append([oct8,8])

  req_count.sort()

  req_octant_ranks=[0,0,0,0,0,0,0,0]

  req_octant_ranks[req_count[0][1]-1]=8   
  req_octant_ranks[req_count[1][1]-1]=7
  req_octant_ranks[req_count[2][1]-1]=6
  req_octant_ranks[req_count[3][1]-1]=5
  req_octant_ranks[req_count[4][1]-1]=4
  req_octant_ranks[req_count[5][1]-1]=3
  req_octant_ranks[req_count[6][1]-1]=2
  req_octant_ranks[req_count[7][1]-1]=1

  octant_rank.append(req_octant_ranks)
  octant_rank_id_name=[]
  octant_rank_id_name=[]

  if req_octant_ranks[0]==1:
   octant_rank_id_name=[1,"Internal outward Interaction"]
   octant_rank1_id.append(octant_rank_id_name)
   r1c=r1c+1
  elif req_octant_ranks[1]==1:
   octant_rank_id_name=[-1,"External outward Interaction"]
   octant_rank1_id.append(octant_rank_id_name)
   r2c=r2c+1
  elif req_octant_ranks[2]==1:
   octant_rank_id_name=[2,"External Ejection"]
   octant_rank1_id.append(octant_rank_id_name)
   r3c=r3c+1 
  elif req_octant_ranks[3]==1:
   octant_rank_id_name=[-2,"Internal Ejection"]
   octant_rank1_id.append(octant_rank_id_name)
   r4c=r4c+1 
  elif req_octant_ranks[4]==1:
   octant_rank_id_name=[3,"External inward Interaction"]
   octant_rank1_id.append(octant_rank_id_name)
   r5c=r5c+1 
  elif req_octant_ranks[5]==1:
   octant_rank_id_name=[-3,"Internal inward Interaction"]
   octant_rank1_id.append(octant_rank_id_name)
   r6c=r6c+1 
  elif req_octant_ranks[6]==1:
   octant_rank_id_name=[4,"Internal Sweep"]
   octant_rank1_id.append(octant_rank_id_name)
   r7c=r7c+1 
  elif req_octant_ranks[7]==1:
   octant_rank_id_name=[-4,"External Sweep"]
   octant_rank1_id.append(octant_rank_id_name)
   r8c=r8c+1 
 

  oct1=0
  oct2=0
  oct3=0
  oct4=0
  oct5=0
  oct6=0
  oct7=0
  oct8=0
 try:
  from openpyxl import Workbook 
  book=Workbook()
  sheet= book.active    
  rows=[] 
  rows.append(["Time",'U','V','W','Uavg','Vavg','Wavg',"U'=U-Uavg","V'=V-Vavg","W'=W-Wavg","Octant","","OctantID","+1","-1","+2","-2","+3","-3","+4","-4","octant_rank of 1","octant_rank of -1","octant_rank of 2","octant_rank of -2","octant_rank of 3","octant_rank of -3","octant_rank of 4","octant_rank of -4","octant_rank1 Octant ID","octant_rank1 Octant Name"])
  j=0
  a=0
  b=0
  for x in range(n-2):
   if(x==0):
    rows.append([time[x],u[x],v[x],w[x],u_avg,v_avg,w_avg,ud[x],vd[x],wd[x],oct[x],"","Overall count",count1,count2,count3,count4,count5,count6,count7,count8,octant_rank[x][0],octant_rank[x][1],octant_rank[x][2],octant_rank[x][3],octant_rank[x][4],octant_rank[x][5],octant_rank[x][6],octant_rank[x][7],octant_rank1_id[x][0],octant_rank1_id[x][1]])
   elif(x==1):
    s="mod "+str(mod)		
    rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"User input",s,"","","","","","","",""])
   elif(x>=2 and x<2+m):
    if(x==1+m):# it will work only if x=1+m
     z=j*mod 	
     y=(j+1)*mod
     s=str(z)+"-"+str(n-2)		 
     rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"",s,octant1[x-2],octant2[x-2],octant3[x-2],octant4[x-2],octant5[x-2],octant6[x-2],octant7[x-2],octant8[x-2],octant_rank[x-1][0],octant_rank[x-1][1],octant_rank[x-1][2],octant_rank[x-1][3],octant_rank[x-1][4],octant_rank[x-1][5],octant_rank[x-1][6],octant_rank[x-1][7],octant_rank1_id[x-1][0],octant_rank1_id[x-1][1]])	 
     j=j+1
    else:
     z=j*mod	
     y=(j+1)*mod-1
     s=str(z)+"-"+str(y)		 
     rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"",s,octant1[x-2],octant2[x-2],octant3[x-2],octant4[x-2],octant5[x-2],octant6[x-2],octant7[x-2],octant8[x-2],octant_rank[x-1][0],octant_rank[x-1][1],octant_rank[x-1][2],octant_rank[x-1][3],octant_rank[x-1][4],octant_rank[x-1][5],octant_rank[x-1][6],octant_rank[x-1][7],octant_rank1_id[x-1][0],octant_rank1_id[x-1][1]])	 
     j=j+1
   elif x==m+5:
    rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","","Octant ID","Octant Name","Count of octant_rank 1 Mod values","","","",""])
   elif x==m+6:
    rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","","1","Internal Outward Interaction",r1c,"","","",""])
   elif x==m+7:
    rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","","-1","External Outward Interaction",r2c,"","","",""])
   elif x==m+8:
    rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","","2","External Ejection",r3c,"","","",""])
   elif x==m+9:
    rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","","-2","Internal Ejection",r4c,"","","",""])
   elif x==m+10:
    rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","","3","External inward Interaction",r5c,"","","",""])
   elif x==m+11:
    rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","","-3","Internal inward Interaction",r6c,"","","",""])
   elif x==m+12:
    rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","","4","Internal Sweep",r7c,"","","",""]) 
   elif x==m+13:
    rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","","-4","External Sweep",r8c,"","","",""])     
   else:
    rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","","","","","","","",""])

  for i in rows:
   sheet.append(i)
  book.save("octant_output_octant_ranking_excel.xlsx")
        

    
  octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
 except:
  print("An exception occurred in printing workbook")
###Code
except:
  print("An exception occurred")
  
from platform import python_version
ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


mod=5000 
octant_range_names(mod)


#This shall be the last lines of the code.  
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
