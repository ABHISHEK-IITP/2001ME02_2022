
from datetime import datetime
start_time = datetime.now()

#Help
def scorecard():
	pass


###Code
####importing required libraries
import openpyxl
import pandas as pd
import os

#reading input files
india_inn = open("india_inns2.txt","r+") #india batting
pak_inn = open("pak_inns1.txt","r+") #pakistan batting
Playing_XI = open("teams.txt","r+")
input_team = Playing_XI.readlines()

pak_team = input_team[0]
pak_players = pak_team[23:-1:].split(",")

ind_team = input_team[2]
ind_players = ind_team[20:-1:].split(",")


lst_ind=india_inn.readlines() #124
for i in lst_ind:
    if i=='\n':
        lst_ind.remove(i)
      

lst_pak=pak_inn.readlines() #123
for i in lst_pak:
    if i=='\n':
        lst_pak.remove(i)

wb = openpyxl.Workbook()
sheet = wb.active

# batting [runs,ball,4s,6s,sr]
# bowling [over,medan,runs,Wickets, NB, WD, ECO]
#declaring required variables
FOW_ind=0
FOW_pak=0
Pak_out_count={}
ind_bowlers={}
ind_bats={}

pak_bats={}
pak_bowlers={}
pak_byes=0
Pak_bowlers_score=0

########Pakistan Innings####################
for l in lst_pak:
    x=l.index(".")
    pak_overs_count=l[0:x+2]
    temp=l[x+2::].split(",")
    c_ball=temp[0].split("to") #0 2

    if f"{c_ball[0].strip()}" not in ind_bowlers.keys() :
        ind_bowlers[f"{c_ball[0].strip()}"]=[1,0,0,0,0,0,0]   #[over0,medan1,runs2,Wickets3, NB4, WD5, ECO6]
    elif "wide" in temp[1]:
        pass
    elif "byes" in temp[1]:                 #defining scores of byes
        if "FOUR" in temp[2]:
            pak_byes+=4
            ind_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "1 run" in temp[2]:
            pak_byes+=1
            ind_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "2 runs" in temp[2]:
            pak_byes+=2
            ind_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "3 runs" in temp[2]:
            pak_byes+=3
            ind_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "4 runs" in temp[2]:
            pak_byes+=4
            ind_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "5 runs" in temp[2]:
            pak_byes+=5
            ind_bowlers[f"{c_ball[0].strip()}"][0]+=1

    else:
        ind_bowlers[f"{c_ball[0].strip()}"][0]+=1
    
    if f"{c_ball[1].strip()}" not in pak_bats.keys() and temp[1]!="wide":
        pak_bats[f"{c_ball[1].strip()}"]=[0,1,0,0,0] #[runs,ball,4s,6s,sr]
    elif "wide" in temp[1] :
        pass
    else:
        pak_bats[f"{c_ball[1].strip()}"][1]+=1
    

    if "out" in temp[1]:                           #updating scoresheet when out
        ind_bowlers[f"{c_ball[0].strip()}"][3]+=1
        if "Bowled" in temp[1].split("!!")[0]:
            Pak_out_count[f"{c_ball[1].strip()}"]=("b" + c_ball[0])
        elif "Caught" in temp[1].split("!!")[0]:
            w=(temp[1].split("!!")[0]).split("by")
            Pak_out_count[f"{c_ball[1].strip()}"]=("c" + w[1] +" b " + c_ball[0])
        elif "Lbw" in temp[1].split("!!")[0]:
            Pak_out_count[f"{c_ball[1].strip()}"]=("lbw  b "+c_ball[0])

    
       #updating scoresheet when run made by bat
    if "no run" in temp[1] or "out" in temp[1] :
        ind_bowlers[f"{c_ball[0].strip()}"][2]+=0
        pak_bats[f"{c_ball[1].strip()}"][0]+=0
    elif "1 run" in temp[1]:
        ind_bowlers[f"{c_ball[0].strip()}"][2]+=1
        pak_bats[f"{c_ball[1].strip()}"][0]+=1
    elif "2 runs" in temp[1]:
        ind_bowlers[f"{c_ball[0].strip()}"][2]+=2
        pak_bats[f"{c_ball[1].strip()}"][0]+=2
    elif "3 runs" in temp[1]:
        ind_bowlers[f"{c_ball[0].strip()}"][2]+=3
        pak_bats[f"{c_ball[1].strip()}"][0]+=3
    elif "4 runs" in temp[1]:
        ind_bowlers[f"{c_ball[0].strip()}"][2]+=4
        pak_bats[f"{c_ball[1].strip()}"][0]+=4
    elif "FOUR" in temp[1]:
        ind_bowlers[f"{c_ball[0].strip()}"][2]+=4
        pak_bats[f"{c_ball[1].strip()}"][0]+=4
        pak_bats[f"{c_ball[1].strip()}"][2]+=1
    elif "SIX" in temp[1]:
        ind_bowlers[f"{c_ball[0].strip()}"][2]+=6
        pak_bats[f"{c_ball[1].strip()}"][0]+=6
        pak_bats[f"{c_ball[1].strip()}"][3]+=1
    elif "wide" in temp[1]:                   #updating scoresheet when wide 
        if "wides" in temp[1]:
            # print(temp[1][1])
            ind_bowlers[f"{c_ball[0].strip()}"][2]+=int(temp[1][1])
            ind_bowlers[f"{c_ball[0].strip()}"][5]+=int(temp[1][1])
        else:
            ind_bowlers[f"{c_ball[0].strip()}"][2]+=1
            ind_bowlers[f"{c_ball[0].strip()}"][5]+=1

for val in pak_bats.values():
    val[-1]=round((val[0]/val[1])*100 , 2)


############# india innings ############## 
ind_bowlers_score=0
ind_byes=0

out_ind_bat={}
for l in lst_ind:
    x=l.index(".")
    over_ind=l[0:x+2]

    temp=l[x+2::].split(",")

    c_ball=temp[0].split("to") #0 2
    if f"{c_ball[0].strip()}" not in pak_bowlers.keys() :
        pak_bowlers[f"{c_ball[0].strip()}"]=[1,0,0,0,0,0,0]   #[over0,medan1,runs2,Wickets3, NB4, WD5, ECO6]
    elif "wide" in temp[1]:
        pass
    elif "byes" in temp[1]:         #updating scoresheet after byes
        if "FOUR" in temp[2]:
            ind_byes+=4
            pak_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "1" in temp[2]:
            ind_byes+=1
            pak_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "2" in temp[2]:
            ind_byes+=2
            pak_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "3" in temp[2]:
            ind_byes+=3
            pak_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "4" in temp[2]:
            ind_byes+=4
            pak_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "5" in temp[2]:
            ind_byes+=5
            pak_bowlers[f"{c_ball[0].strip()}"][0]+=1
    else:
        pak_bowlers[f"{c_ball[0].strip()}"][0]+=1
    
    if f"{c_ball[1].strip()}" not in ind_bats.keys() and temp[1]!="wide":
        ind_bats[f"{c_ball[1].strip()}"]=[0,1,0,0,0] #[runs,ball,4s,6s,sr]
    elif "wide" in temp[1] :
        pass
    else:
        ind_bats[f"{c_ball[1].strip()}"][1]+=1
    

    if "out" in temp[1]:                    #updating scoresheet after out
        pak_bowlers[f"{c_ball[0].strip()}"][3]+=1
        
        if "Bowled" in temp[1].split("!!")[0]:
            out_ind_bat[f"{c_ball[1].strip()}"]=("b" + c_ball[0])
        elif "Caught" in temp[1].split("!!")[0]:
            w=(temp[1].split("!!")[0]).split("by")
            out_ind_bat[f"{c_ball[1].strip()}"]=("c" + w[1] +" b " + c_ball[0])
        elif "Lbw" in temp[1].split("!!")[0]:
            out_ind_bat[f"{c_ball[1].strip()}"]=("lbw  b "+c_ball[0])

    
    
    if "no run" in temp[1] or "out" in temp[1] :    #updating scoresheet after runs by bat
        pak_bowlers[f"{c_ball[0].strip()}"][2]+=0
        ind_bats[f"{c_ball[1].strip()}"][0]+=0
    elif "1 run" in temp[1]:
        pak_bowlers[f"{c_ball[0].strip()}"][2]+=1
        ind_bats[f"{c_ball[1].strip()}"][0]+=1
    elif "2 runs" in temp[1]:
        pak_bowlers[f"{c_ball[0].strip()}"][2]+=2
        ind_bats[f"{c_ball[1].strip()}"][0]+=2
    elif "3 runs" in temp[1]:
        pak_bowlers[f"{c_ball[0].strip()}"][2]+=3
        ind_bats[f"{c_ball[1].strip()}"][0]+=3
    elif "4 runs" in temp[1]:
        pak_bowlers[f"{c_ball[0].strip()}"][2]+=4
        ind_bats[f"{c_ball[1].strip()}"][0]+=4
    elif "FOUR" in temp[1]:
        pak_bowlers[f"{c_ball[0].strip()}"][2]+=4
        ind_bats[f"{c_ball[1].strip()}"][0]+=4
        ind_bats[f"{c_ball[1].strip()}"][2]+=1
    elif "SIX" in temp[1]:
        pak_bowlers[f"{c_ball[0].strip()}"][2]+=6
        ind_bats[f"{c_ball[1].strip()}"][0]+=6
        ind_bats[f"{c_ball[1].strip()}"][3]+=1
    elif "wide" in temp[1]:
        if "wides" in temp[1]:
            pak_bowlers[f"{c_ball[0].strip()}"][2]+=int(temp[1][1])
            pak_bowlers[f"{c_ball[0].strip()}"][5]+=int(temp[1][1])
        else:
            pak_bowlers[f"{c_ball[0].strip()}"][2]+=1
            pak_bowlers[f"{c_ball[0].strip()}"][5]+=1


for val in ind_bats.values():
    val[-1]=round((val[0]/val[1])*100 , 2)

for val in pak_bats.values():
    val[-1]=round((val[0]/val[1])*100 , 2)

for val in ind_bowlers.values():
    if val[0]%6==0:
        val[0] = val[0]//6
    else:
        val[0] = (val[0]//6) + (val[0]%6)/10

for val in pak_bowlers.values():
    if val[0]%6==0:
        val[0] = val[0]//6
    else:
        val[0] = (val[0]//6) + (val[0]%6)/10

for val in ind_bowlers.values(): #economy
    x=str(val[0])
    if "." in x:
        balls = int(x[0])*6 + int(x[2])
        val[-1]=round((val[2]/balls)*6,1)
    else:
        val[-1] = round((val[2]/val[0]) ,1) 


for val in pak_bowlers.values(): #economy
    x=str(val[0])
    if "." in x:
        balls = int(x[0])*6 + int(x[2])
        val[-1]=round((val[2]/balls)*6,1)
    else:
        val[-1] = round((val[2]/val[0]) ,1)

Fall_of_Wickets_paki="15-1(Babar Azam,2.4), 42-2(Fakhar Zaman,5.5), 87-3(Iftikhar Ahmed,12.1), 96-4(Rizwan,14.1), 97-5(Khushdil,14.3), 112-6(Asif Ali,16.3), 114-7(Mohammad Nawaz,17.1), 128-8(Shadab Khan,18.2), 128-9(Naseem Shah,18.3), 147-10(Dahani,19.5)"
Fall_of_Wickets_india = "1-1(Rahul,0.2), 50-2(Rohit,8.0), 53-3(Kohli,9.1), 89-4(Suryakumar Yadav,14.2), 136-5(Jadeja,19.1)"

# pakistan batting
Pak_batters_name=[]
for key in pak_bats.keys():
    Pak_batters_name.append(key)


for i in range(len(pak_bats)):
    sheet.cell(5+i,1).value = Pak_batters_name[i]
    sheet.cell(5+i,5).value = pak_bats[Pak_batters_name[i]][0]
    sheet.cell(5+i,6).value = pak_bats[Pak_batters_name[i]][1]
    sheet.cell(5+i,7).value = pak_bats[Pak_batters_name[i]][2]
    sheet.cell(5+i,8).value = pak_bats[Pak_batters_name[i]][3]
    sheet.cell(5+i,9).value = pak_bats[Pak_batters_name[i]][4]
    if Pak_batters_name[i] not in Pak_out_count:
        sheet.cell(5+i,3).value = "not out"
    else:
        sheet.cell(5+i,3).value=Pak_out_count[Pak_batters_name[i]]


extra_data_paki = "5 (b 1, lb 0, w 4, nb 0, p 0)"
extra_data_india = "14 (b 0, lb 5, w 9, nb 0, p 0)"

sheet.cell(3,1).value = "Batter"
sheet["E3"] = "Runs"
sheet["F3"] = "Balls"
sheet["G3"] = " 4s "
sheet["H3"] = " 6s "
sheet["I3"] = "  SR  "

# Pakistan bowling

sheet["A21"] = "Bowler"
sheet["C21"] = "Over"
sheet["D21"] = "Maiden"
sheet["E21"] = "Runs"
sheet["F21"] = "Wicket"
sheet["G21"] = "No Ball"
sheet["H21"] = "Wide"
sheet["I21"] = "Economy"

Pak_bowlers_name=[]
for key in pak_bowlers.keys():
    Pak_bowlers_name.append(key)

for i in range(len(pak_bowlers)):
    sheet.cell(47+i,1).value = Pak_bowlers_name[i]
    sheet.cell(47+i,3).value = pak_bowlers[Pak_bowlers_name[i]][0]
    sheet.cell(47+i,4).value = pak_bowlers[Pak_bowlers_name[i]][1]
    sheet.cell(47+i,5).value = pak_bowlers[Pak_bowlers_name[i]][2]
    sheet.cell(47+i,6).value = pak_bowlers[Pak_bowlers_name[i]][3]
    sheet.cell(47+i,7).value = pak_bowlers[Pak_bowlers_name[i]][4]
    sheet.cell(47+i,8).value = pak_bowlers[Pak_bowlers_name[i]][5]
    sheet.cell(47+i,9).value = pak_bowlers[Pak_bowlers_name[i]][6]
    Pak_bowlers_score+=pak_bowlers[Pak_bowlers_name[i]][2]
    FOW_ind+=pak_bowlers[Pak_bowlers_name[i]][3]

# india batting
sheet.cell(14+len(pak_bats)+len(pak_bowlers),1).value = " INDIA"
sheet.cell(14+len(pak_bats)+len(pak_bowlers),2).value = " INNINGS"

Ind_batters_name=[]
for key in ind_bats.keys():
    Ind_batters_name.append(key)


for i in range(len(ind_bats)):
    sheet.cell(34+i,1).value = Ind_batters_name[i]
    sheet.cell(34+i,5).value = ind_bats[Ind_batters_name[i]][0]
    sheet.cell(34+i,6).value = ind_bats[Ind_batters_name[i]][1]
    sheet.cell(34+i,7).value = ind_bats[Ind_batters_name[i]][2]
    sheet.cell(34+i,8).value = ind_bats[Ind_batters_name[i]][3]
    sheet.cell(34+i,9).value = ind_bats[Ind_batters_name[i]][4]

    if Ind_batters_name[i] not in out_ind_bat:
        sheet.cell(34+i,3).value = "not out"
    else:
        sheet.cell(34+i,3).value=out_ind_bat[Ind_batters_name[i]]

sheet["A32"] = "Batter"
sheet["E32"] = "Runs"
sheet["F32"] = "Balls"
sheet["G32"] = " 4s "
sheet["H32"] = " 6s "
sheet["I32"] = "  SR  "

# india bowling

sheet["A46"] = "Bowler"
sheet["C46"] = "Over"
sheet["D46"] = "Maiden"
sheet["E46"] = "Runs"
sheet["F46"] = "Wicket"
sheet["G46"] = "No Ball"
sheet["H46"] = "Wide"
sheet["I46"] = "Economy"

Ind_bowlers_name=[]
for key in ind_bowlers.keys():
    Ind_bowlers_name.append(key)

for i in range(len(ind_bowlers)):

    sheet.cell(22+i,1).value = Ind_bowlers_name[i]
    sheet.cell(22+i,3).value = ind_bowlers[Ind_bowlers_name[i]][0]
    sheet.cell(22+i,4).value = ind_bowlers[Ind_bowlers_name[i]][1]
    sheet.cell(22+i,5).value = ind_bowlers[Ind_bowlers_name[i]][2]
    sheet.cell(22+i,6).value = ind_bowlers[Ind_bowlers_name[i]][3]
    sheet.cell(22+i,7).value = ind_bowlers[Ind_bowlers_name[i]][4]
    sheet.cell(22+i,8).value = ind_bowlers[Ind_bowlers_name[i]][5]
    sheet.cell(22+i,9).value = ind_bowlers[Ind_bowlers_name[i]][6]
    ind_bowlers_score+=ind_bowlers[Ind_bowlers_name[i]][2]
    FOW_pak+=ind_bowlers[Ind_bowlers_name[i]][3]

ind_total_score=ind_bowlers_score+pak_byes +1
pak_total_score = Pak_bowlers_score+ind_byes -1
# print(over_ind)
# print(pak_overs_count)
sheet["H30"] = " "+str(ind_total_score) +" - " + str(FOW_ind)
sheet["I30"] = "(" +str(over_ind)+ ")" +"overs"
Eone=" "+str(pak_total_score) +" - " + str(FOW_pak)
Fone = "(" +str(pak_overs_count) + ")" +"overs"

sheet["A17"]="Extra"
sheet["E17"]=extra_data_paki
sheet["A18"] = "Total"
sheet["E18"] = str(pak_total_score)+" ("+ str(FOW_pak) + " wkt, "+ str(pak_overs_count) + " Ov )"
sheet["A19"] = "Fall of Wickets"
sheet["E19"] = Fall_of_Wickets_paki

sheet["A42"] = "Extra"
sheet["E42"] = extra_data_india
sheet["A43"] = "Total"
sheet["E43"] = str(ind_total_score)+" ("+ str(FOW_ind) + " wkt, "+ str(over_ind) + " Ov )"
sheet["A44"] = "Fall of Wickets"
sheet["E44"] = Fall_of_Wickets_india

wb.save("Scoreboard.xlsx")

df = pd.read_excel('Scoreboard.xlsx')

df = df.set_axis(['PAKISTAN', ' Innings'] + [" "," ","",""," ",Eone ,Fone], axis='columns')

df.to_csv('Scorecard.csv',index=False)

import os

os.path.exists("Scoreboard.xlsx") 
os.remove("Scoreboard.xlsx") # deleting output excel





from platform import python_version
ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


scorecard()






#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
