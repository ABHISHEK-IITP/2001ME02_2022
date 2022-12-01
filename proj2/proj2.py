
from datetime import datetime
import streamlit as st
st.title("Hello")
start_time = datetime.now()



#Help
def proj_octant_gui():

	pass
	############ importing libraries rquired
	from openpyxl.styles.borders import Border, Side
	import pandas as pd
	import os.path
	import pathlib
	import shutil
	from io import BytesIO
	import datetime
	now = datetime.datetime.now()

	###### Title of the page
	st.title("Project 2 (CS384)")
	


	######Borders for sheet
	make_border = Border(left=Side(style='thin'), right=Side(style='thin'),top=Side(style='thin'), bottom=Side(style='thin'))
	comp_head=['T','U','V','W','U Avg','V Avg','W Avg',r"U'=U-U avg",r"V'=V-V avg",r"W'=W-W avg",'Octant']

	########## Defining function to find octants
	def oct_count(a,b,c):                                      
		if(a>0 and b>0 and c>0):    
			return 1
		elif(a>0 and b>0 and c<0):
			return -1
		elif(a<0 and b>0 and c>0):
			return 2
		elif(a<0 and b>0 and c<0):
			return -2
		elif(a<0 and b<0 and c>0):
			return 3
		elif(a<0 and b<0 and c<0):
			return -3
		elif(a>0 and b<0 and c>0):
			return 4
		elif(a>0 and b<0 and c<0):
			return -4

	######### Defining function for single input file for octant analysis
	def octant_analysis_single(mod,file):
		from pandas import read_excel
		import openpyxl
		from openpyxl import workbook,load_workbook
		from openpyxl import Workbook
		from openpyxl.styles import PatternFill
		from itertools import repeat

		##########Defining working file function
		def workingFile(s,mod):							
			inputFilePath=s									
			df=read_excel(inputFilePath)									
			wb=Workbook()											
			sheet=wb.active													
			
			for i in range(11):										
				sheet.cell(row=2,column=i+1).value=comp_head[i]	
			octant=[]

			u_avg=df['U'].mean()                                                 
			v_avg=df['V'].mean()
			w_avg=df['W'].mean()
			sheet.cell(row=1,column=14).value='Overall Octant Count'
			sheet.cell(row=1,column=45).value='Longest Subsequence Length'
			sheet.cell(row=1,column=49).value='Longest Subsquence Length with Range'
			sheet['E3'] = u_avg
			sheet['F3'] = v_avg
			sheet['G3'] = w_avg


			for i in df.index:
				sheet.cell(row=i+3,column=1).value=df['T'][i]
				sheet.cell(row=i+3,column=2).value=df['U'][i]
				sheet.cell(row=i+3,column=3).value=df['V'][i]
				sheet.cell(row=i+3,column=4).value=df['W'][i]
				sheet.cell(row=i+3,column=8).value=round(df['U'][i]-u_avg,3)
				sheet.cell(row=i+3,column=9).value=round(df['V'][i]-v_avg,3)
				sheet.cell(row=i+3,column=10).value=round(df['W'][i]-w_avg,3)
				sheet.cell(row=i+3,column=11).value=oct_count(df['U'][i]-u_avg,df['V'][i]-v_avg,df['W'][i]-w_avg)
				octant.append(oct_count(df['U'][i]-u_avg,df['V'][i]-v_avg,df['W'][i]-w_avg))
			
			######## Defining function to name the octant range
			def name_oct_range(mod):
				octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}

				# defining dictionary for mapping
				map_dict={}                            

				# Creating dictionary with opposite key value pair than 'map_dict'                              
				opp_map_dict={}                                                      
				
				for i in range(4):                                            
					map_dict[i+1]=2*i+1-1                                           
					map_dict[-(i+1)]=2*(i+1)-1
					opp_map_dict[2*i+1-1]=i+1
					opp_map_dict[2*(i+1)-1]=-(i+1)

				# Function to find the rank list from count values of all octamts	
				def rank_list(comp_head):                                     
					temp_comp_head=comp_head.copy()
					temp_comp_head.sort(reverse=True)
					res=[]

					for i in comp_head:
						for j in range(8):
							if(i==temp_comp_head[j]):
								res.append(j+1)
								break
					
					return res                                                  
				
				# Finding the octant which has rank 1 in the given rank list
				def find_rank1(comp_head):                                         
					for i in range(8):
						if(comp_head[i]==1):
							return opp_map_dict[i]

				# Finding the count of rank 1 in the rank 1 mod values of octant x
				def count_rank1(comp_head,x):                                         
					add=0
					for i in comp_head:
						if(x==i):
							add+=1
					return add                                                  
				
				######### Defining Matrix to store rank list 
				matrix_rank=[] 

				###### Defining List to store the octs which have rank 1                                               
				list_rank1=[] 
				try:                                                  
					sheet=wb.active
				except:
					print("sheet activation error")	

				sheet['M4']='Mod '+str(mod)                                           

				oct_range_matrix=[]   
                                                  
				count1=[0]*9                                                     

				count1[0]='Octant ID'                                           

				for i in range(0,4):
					count1[2*i+1]=(i+1)
					count1[2*(i+1)]=-(i+1)

				####### Appending header list in matrix	
				oct_range_matrix.append(count1)          

				####### Writing header list in workbook                                 
				for i in range(13,22):                                          
					sheet.cell(row=3,column=i+1).value=count1[i-13]
					sheet.cell(row=3,column=i+1).border=make_border
					if(i>13):
						sheet.cell(row=3,column=i+9).value='Rank Octant '+str(count1[i-13])
						sheet.cell(row=3,column=i+9).border=make_border
				sheet.cell(row=3,column=31).value='Rank1 Octant ID'
				sheet.cell(row=3,column=32).value='Rank1 Octant Name'
				sheet.cell(row=3,column=31).border=make_border
				sheet.cell(row=3,column=32).border=make_border

			
				count1=[0]*9                                                     
				for i in octant:                                                
					if(i==1):
						count1[1]=count1[1]+1
					elif(i==-1):
						count1[2]=count1[2]+1
					elif(i==2):
						count1[3]=count1[3]+1
					elif(i==-2):
						count1[4]=count1[4]+1
					elif(i==3):
						count1[5]=count1[5]+1
					elif(i==-3):
						count1[6]=count1[6]+1
					elif(i==4):
						count1[7]=count1[7]+1
					elif(i==-4):
						count1[8]=count1[8]+1

				CodeYelow = "00FFFF00"

				count1[0]='Overall Count'  
				##### Appending octant range matrix by count 1                                      
				oct_range_matrix.append(count1)          

				# Writing overall count in worksheet                                 
				for i in range(13,22):                                         
					sheet.cell(row=4,column=i+1).value=count1[i-13]
					sheet.cell(row=4,column=i+1).border=make_border
				
				count1.pop(0)                                              
				rank_rank=rank_list(count1)             

				# Finding the rank 1 octant and appending in list_rank1                      
				list_rank1.append(find_rank1(rank_rank))                         
				matrix_rank.append(rank_rank)            
                   
				for i in range(8):                                              
					sheet.cell(row=4,column=23+i).value=matrix_rank[0][i]
					sheet.cell(row=4,column=23+i).border=make_border
					if(matrix_rank[0][i]==1):
						sheet.cell(row=4,column=23+i).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
				sheet.cell(row=4,column=31).value=list_rank1[0]
				sheet.cell(row=4,column=32).value=octant_name_id_mapping[str(list_rank1[0])]
				sheet.cell(row=4,column=31).border=make_border
				sheet.cell(row=4,column=32).border=make_border
										
				
				####### Finding the number of points given in the input
				m=len(octant)                                              
				count1=[0]*9          

	                                        
				k=0                                                   
				j=4               

				##### Counting number of values in different octs in mod range                                              
				for i in octant:                                               
					if(i==1):
						count1[1]=count1[1]+1
					elif(i==-1):
						count1[2]=count1[2]+1
					elif(i==2):
						count1[3]=count1[3]+1
					elif(i==-2):
						count1[4]=count1[4]+1
					elif(i==3):
						count1[5]=count1[5]+1
					elif(i==-3):
						count1[6]=count1[6]+1
					elif(i==4):
						count1[7]=count1[7]+1
					elif(i==-4):
						count1[8]=count1[8]+1
					k=k+1                                                     
					if(k%mod==1):                                              
						count1[0]=str(k-1)+'-'                       
					elif(k%mod==0 or k==m):
						count1[0]=count1[0]+str(k-1)    

						##### Writing the columns of rank, rank1 and octant_name 
						for i in range(13,22):                                
							sheet.cell(row=j+1,column=i+1).value=count1[i-13]
							sheet.cell(row=j+1,column=i+1).border=make_border
						count1.pop(0)                                         
						rank_rank=rank_list(count1)                           
						list_rank1.append(find_rank1(rank_rank))                
						matrix_rank.append(rank_rank)                            
						
					
						for i in range(8):                                              
							sheet.cell(row=j+1,column=23+i).value=matrix_rank[j-3][i]
							sheet.cell(row=j+1,column=23+i).border=make_border
							if(matrix_rank[j-3][i]==1):
								sheet.cell(row=j+1,column=23+i).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
						
						sheet.cell(row=j+1,column=31).value=list_rank1[j-3]
						sheet.cell(row=j+1,column=32).value=octant_name_id_mapping[str(list_rank1[j-3])]
						sheet.cell(row=j+1,column=31).border=make_border
						sheet.cell(row=j+1,column=32).border=make_border
						
						j=j+1
						oct_range_matrix.append(count1)
						count1=[0]*9                                
														
				list_rank1.pop(0)                                         
				j+=4
				sheet.cell(row=j,column=29).value='Octant ID'        
				sheet.cell(row=j,column=30).value='Octant Name'
				sheet.cell(row=j,column=31).value='Count of Rank 1 Mod Values'
				sheet.cell(row=j,column=29).border=make_border                 
				sheet.cell(row=j,column=30).border=make_border
				sheet.cell(row=j,column=31).border=make_border
				
				# Writing the table of count of rank1 mod values
				for i in range(8):                              
					sheet.cell(row=j+1+i,column=29).value=opp_map_dict[i]
					sheet.cell(row=j+1+i,column=30).value=octant_name_id_mapping[str(opp_map_dict[i])]
					sheet.cell(row=j+1+i,column=31).value=count_rank1(list_rank1,opp_map_dict[i])
					sheet.cell(row=j+1+i,column=29).border=make_border
					sheet.cell(row=j+1+i,column=30).border=make_border
					sheet.cell(row=j+1+i,column=31).border=make_border
			
			def oct_lsubs_count_with_range():
				
				# Header list
				r=['Count','Longest Subsequence Length','Count']

				# Writing header of table to worksheet        
				for i in range(3):                                     
					sheet.cell(row=3,column=45+i).value=r[i] 
					sheet.cell(row=3,column=45+i).border=make_border    


				octs=[]

				for i in range(2,10,2):                                    
					sheet.cell(row=i+2,column=45).value=i//2
					octs.append(i//2)
					sheet.cell(row=i+3,column=45).value=-(i//2) 
					octs.append(-i//2)
					sheet.cell(row=i+2,column=45).border=make_border
					sheet.cell(row=i+3,column=45).border=make_border                                
				
				
				map_dict={}                                              
				for i in range(0,4):                                            
					map_dict[i+1]=2*i+1-1
					map_dict[-(i+1)]=2*(i+1)-1

				
				count1=[0]*8                                             
				long_len=[0]*8                                    
				back=octant[0]

			
				l=1                                                           
				n=len(octant)

				temp_max=[0]     

				ranges= [[] for x in repeat(None, 8)]            

				
				for i in range(1,n+1):  
					             
					if(i==n):                                        
						if(long_len[map_dict[back]]<l):                        
							long_len[map_dict[back]]=l
							count1[map_dict[back]]=1

							# Writing ending range in temp
							temp_max.append(df['T'][i-1])            
							ranges[map_dict[back]].clear()         
							ranges[map_dict[back]].append(temp_max)           
						elif(long_len[map_dict[back]]==l):
							count1[map_dict[back]]+=1
							temp_max.append(df['T'][i-1])
							ranges[map_dict[back]].append(temp_max) 	               
					elif(back==octant[i]):                                      
						l+=1

					else:                                                      
						if(long_len[map_dict[back]]<l):
							long_len[map_dict[back]]=l
							count1[map_dict[back]]=1
							ranges[map_dict[back]].clear()               
							temp_max.append(df['T'][i-1])        
							ranges[map_dict[back]].append(temp_max)                 
						elif(long_len[map_dict[back]]==l):
							count1[map_dict[back]]+=1
							temp_max.append(df['T'][i-1])
							ranges[map_dict[back]].append(temp_max)             
						temp_max=[df['T'][i]]                     
						l=1
						back=octant[i]                            

				# Writing the number and length of longest subsequence in table
				for i in range(2,10):                                    
					sheet.cell(row=i+2,column=46).value=long_len[i-2]
					sheet.cell(row=i+2,column=47).value=count1[i-2]
					sheet.cell(row=i+2,column=46).border=make_border
					sheet.cell(row=i+2,column=47).border=make_border
				k=2                                                        
				sheet.cell(row=k+1,column=49).value='Octant ###'                  
				sheet.cell(row=k+1,column=50).value='Longest Subsequence Length'
				sheet.cell(row=k+1,column=51).value='Count'
				sheet.cell(row=k+1,column=49).border=make_border
				sheet.cell(row=k+1,column=50).border=make_border
				sheet.cell(row=k+1,column=51).border=make_border
				
				k+=2
				for i in range(8):
					sheet.cell(row=k,column=49).value=octs[i]            
					sheet.cell(row=k,column=50).value=long_len[i]
					sheet.cell(row=k,column=51).value=count1[i]
					sheet.cell(row=k+1,column=49).value='Time'             
					sheet.cell(row=k+1,column=50).value='From'
					sheet.cell(row=k+1,column=51).value='To'
					sheet.cell(row=k,column=49).border=make_border      
					sheet.cell(row=k,column=50).border=make_border
					sheet.cell(row=k,column=51).border=make_border
					sheet.cell(row=k+1,column=49).border=make_border
					sheet.cell(row=k+1,column=50).border=make_border
					sheet.cell(row=k+1,column=51).border=make_border
					x=ranges[i]
					k+=2
					for j in x:
						# Writing ranges in worksheet
						sheet.cell(row=k,column=50).value=j[0]     
						sheet.cell(row=k,column=51).value=j[1]
						sheet.cell(row=k,column=49).border=make_border  
						sheet.cell(row=k,column=50).border=make_border
						sheet.cell(row=k,column=51).border=make_border
						k+=1
				
			def octant_transition_count(mod=5000):
				j=1
				n=len(octant)
				sheet.cell(row=j,column=35).value='Overall Transition Count' 
				sheet.cell(row=j+3,column=34).value='From'
				sheet.cell(row=j+1,column=36).value='To'
				j+=2
				
				oct_range_matrix = [ [0]*9 for i in range(9)]                    
				
				# Storing header row and header column in the matrix
				for i in range(0,4):                                        
					oct_range_matrix[0][2*i+1]=(i+1)
					oct_range_matrix[0][2*(i+1)]=-(i+1)
				for i in range(0,9):
					oct_range_matrix[i][0]=oct_range_matrix[0][i]
				oct_range_matrix[0][0]='Octant #'

				# creating dictionary for mapping 
				map_dict={}                                             
				for i in range(0,4):
					map_dict[i+1]=2*i+1
					map_dict[-(i+1)]=2*(i+1)

				# Finding row and column of matrix from transition values
				def find_rc(x,y):                     
					comp_head=[map_dict[x],map_dict[y]]
					return comp_head
				
				def max_ele(comp_head):
					temp_max=comp_head.copy()
					temp_max.pop(0)
					large=0
					for i in temp_max:
						if(large<i):
							large=i
					return large

				back=octant[0]

				# Filling overall transition matrix                                              
				for i in range(1,n):                       
					comp_head=find_rc(back,octant[i])                        
					oct_range_matrix[comp_head[0]][comp_head[1]]+=1
					back=octant[i]

				CodeYelow = "00FFFF00"

				# Writing the overall transition matrix in worksheet
				for i in range(9):                                  
					temp_comp_head=oct_range_matrix[i]
					large=max_ele(temp_comp_head)
					for k in range(13,22):
						sheet.cell(row=j+i,column=k+22).value=oct_range_matrix[i][k-13]
						sheet.cell(row=j+i,column=k+22).border=make_border
						if(i>0 and oct_range_matrix[i][k-13]==large):
							sheet.cell(row=j+i,column=k+22).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
						if(i!=0 and k!=13):
							oct_range_matrix[i][k-13]=0

				# temp-> No. of mod transition tables	
				temp_max=n//mod+1                                              
				j+=1

				# One iteration for each mod transition table
				for t in range(temp_max):                                   
					j+=11
					name1=''

					# Writing Table name in worksheet
					sheet.cell(row=j,column=35).value='Mod Transition Count'     
					sheet.cell(row=j+3,column=34).value='From'
					sheet.cell(row=j+1,column=36).value='To'
					name1=str(t*mod)+'-'
					if((t+1)*mod-1>n-1):
						name1+=str(n-1)
					else:
						name1+=str((t+1)*mod-1)   
					sheet.cell(row=j+1,column=35).value=name1
					j+=2

					# Incrementing matrix cell corresponding to transition values
					for i in range(t*mod,min(n-1,(t+1)*mod)):                 
						comp_head=find_rc(octant[i],octant[i+1])
						oct_range_matrix[comp_head[0]][comp_head[1]]+=1

					# Writing the transition mod matrix in worksheet
					for i in range(0,9):                                     
						temp_comp_head=oct_range_matrix[i]
						if(i>0):
							large=max_ele(temp_comp_head)
						for k in range(13,22):
							sheet.cell(row=j+i,column=k+22).value=oct_range_matrix[i][k-13]
							sheet.cell(row=j+i,column=k+22).border=make_border
							if(i>0 and oct_range_matrix[i][k-13]==large):
								sheet.cell(row=j+i,column=k+22).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
							if(i!=0 and k!=13):

								# Resetting matrix for next mod iteration
								oct_range_matrix[i][k-13]=0                           


			octant_transition_count(mod)
			name_oct_range(mod)
			oct_lsubs_count_with_range()
			s=s.name[:-5]
			path_out='output/'+s+' cm_vel_octant_analysis_mod_'+str(mod)+'.xlsx'
			wb.save(path_out)

			data_1 = BytesIO()
			wb.save(data_1)
			filename = s+ "cm_vel_octant_analysis_mod_"+str(val)+".xlsx" 
			st.text(filename) 
			st.download_button(label="Download File", file_name=s+
							"cm_vel_octant_analysis_mod_"+str(val)+str(now)+".xlsx", data=data_1)
		# a = pathlib.Path(_file_).parent.parent.resolve()
		# input_files = os.path.join(a, "proj2\input")
		# # input_files=os.listdir('input')
		# for i in range(len(input_files)):
		# 	workingFile(mod,file)
		workingFile(file,mod)

	#**********MULTIPLE file*******************

	def octant_analysis_mutiple(mod,path):
		from pandas import read_excel
		import openpyxl
		from openpyxl import workbook,load_workbook
		from openpyxl import Workbook
		from openpyxl.styles import PatternFill
		from itertools import repeat

		def workingFile(s,file,mod):							
			inputFilePath=s									
			df=read_excel(inputFilePath)									
			wb=Workbook()											
			sheet=wb.active													
			
			for i in range(11):										
				sheet.cell(row=2,column=i+1).value=comp_head[i]	
			octant=[]

			u_avg=df['U'].mean()                                                 
			v_avg=df['V'].mean()
			w_avg=df['W'].mean()
			sheet.cell(row=1,column=14).value='Overall Octant Count'
			sheet.cell(row=1,column=45).value='Longest Subsequence Length'
			sheet.cell(row=1,column=49).value='Longest Subsquence Length with Range'
			sheet['E3'] = u_avg
			sheet['F3'] = v_avg
			sheet['G3'] = w_avg


			for i in df.index:
				sheet.cell(row=i+3,column=1).value=df['T'][i]
				sheet.cell(row=i+3,column=2).value=df['U'][i]
				sheet.cell(row=i+3,column=3).value=df['V'][i]
				sheet.cell(row=i+3,column=4).value=df['W'][i]
				sheet.cell(row=i+3,column=8).value=round(df['U'][i]-u_avg,3)
				sheet.cell(row=i+3,column=9).value=round(df['V'][i]-v_avg,3)
				sheet.cell(row=i+3,column=10).value=round(df['W'][i]-w_avg,3)
				sheet.cell(row=i+3,column=11).value=oct_count(df['U'][i]-u_avg,df['V'][i]-v_avg,df['W'][i]-w_avg)
				octant.append(oct_count(df['U'][i]-u_avg,df['V'][i]-v_avg,df['W'][i]-w_avg))
			
			def name_oct_range(mod=5000):
				octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}

				# creating dictionary for mapping 
				map_dict={}                            

				# Creating dictionary with opposite key value pair than 'map_dict'                              
				opp_map_dict={}                                                      
				
				for i in range(4):                                            
					map_dict[i+1]=2*i+1-1                                           
					map_dict[-(i+1)]=2*(i+1)-1
					opp_map_dict[2*i+1-1]=i+1
					opp_map_dict[2*(i+1)-1]=-(i+1)

				# Function to find the rank list from count values of all octs	
				def rank_list(comp_head):                                     
					temp_comp_head=comp_head.copy()
					temp_comp_head.sort(reverse=True)
					res=[]

					for i in comp_head:
						for j in range(8):
							if(i==temp_comp_head[j]):
								res.append(j+1)
								break
			
					return res                                                  
				
				# Finding the octant which has rank 1 in the given rank list
				def find_rank1(comp_head):                                         
					for i in range(8):
						if(comp_head[i]==1):
							return opp_map_dict[i]

				# Finding the count of rank 1 in the rank 1 mod values of octant x
				def count_rank1(comp_head,x):                                         
					add=0
					for i in comp_head:
						if(x==i):
							add+=1
					return add                                                  
				
				# Matrix to store rank list for different mod values
				matrix_rank=[] 

				# List to store the octs which have rank 1 in different mod ranges and overall                                               
				list_rank1=[] 
				try:                                                  
					sheet=wb.active
				except:
					print("sheet activation error")	

				# Putting the string 'User Input' at its specified place
				sheet['M4']='Mod '+str(mod)                                           

				oct_range_matrix=[]   
                                                   
				count1=[0]*9                                                     

				# Storing header list in 'count1' list
				count1[0]='Octant ID'                                           

				for i in range(0,4):
					count1[2*i+1]=(i+1)
					count1[2*(i+1)]=-(i+1)
	
				oct_range_matrix.append(count1)          

	                                 
				for i in range(13,22):                                          
					sheet.cell(row=3,column=i+1).value=count1[i-13]
					sheet.cell(row=3,column=i+1).border=make_border
					if(i>13):
						sheet.cell(row=3,column=i+9).value='Rank Octant '+str(count1[i-13])
						sheet.cell(row=3,column=i+9).border=make_border
				sheet.cell(row=3,column=31).value='Rank1 Octant ID'
				sheet.cell(row=3,column=32).value='Rank1 Octant Name'
				sheet.cell(row=3,column=31).border=make_border
				sheet.cell(row=3,column=32).border=make_border

		
				count1=[0]*9                                                     

	
				for i in octant:                                                
					if(i==1):
						count1[1]=count1[1]+1
					elif(i==-1):
						count1[2]=count1[2]+1
					elif(i==2):
						count1[3]=count1[3]+1
					elif(i==-2):
						count1[4]=count1[4]+1
					elif(i==3):
						count1[5]=count1[5]+1
					elif(i==-3):
						count1[6]=count1[6]+1
					elif(i==4):
						count1[7]=count1[7]+1
					elif(i==-4):
						count1[8]=count1[8]+1

				##### color code(yellow) given to rank 1 cells
				CodeYelow = "00FFFF00"

			
				count1[0]='Overall Count'                                        
				oct_range_matrix.append(count1)          

		                          
				for i in range(13,22):                                         
					sheet.cell(row=4,column=i+1).value=count1[i-13]
					sheet.cell(row=4,column=i+1).border=make_border
				count1.pop(0)                     
                       
				rank_rank=rank_list(count1)                                
				list_rank1.append(find_rank1(rank_rank))                         
				matrix_rank.append(rank_rank)            

				# Writing overall count in worksheet                           
				for i in range(8):                                              
					sheet.cell(row=4,column=23+i).value=matrix_rank[0][i]
					sheet.cell(row=4,column=23+i).border=make_border
					if(matrix_rank[0][i]==1):
						sheet.cell(row=4,column=23+i).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
				sheet.cell(row=4,column=31).value=list_rank1[0]
				sheet.cell(row=4,column=32).value=octant_name_id_mapping[str(list_rank1[0])]
				sheet.cell(row=4,column=31).border=make_border
				sheet.cell(row=4,column=32).border=make_border
										
				
				m=len(octant)    
				# Resetting the values in the list 'count'                                              
				count1=[0]*9          
			                                          
				k=0                               
				j=4               
				# Counting number of values in different mod values                                             
				for i in octant:                                               
					if(i==1):
						count1[1]=count1[1]+1
					elif(i==-1):
						count1[2]=count1[2]+1
					elif(i==2):
						count1[3]=count1[3]+1
					elif(i==-2):
						count1[4]=count1[4]+1
					elif(i==3):
						count1[5]=count1[5]+1
					elif(i==-3):
						count1[6]=count1[6]+1
					elif(i==4):
						count1[7]=count1[7]+1
					elif(i==-4):
						count1[8]=count1[8]+1
					k=k+1                                                     
					if(k%mod==1):                                              
						count1[0]=str(k-1)+'-'                       
					elif(k%mod==0 or k==m):
						count1[0]=count1[0]+str(k-1)    

						# Writing the columns of rank, rank1 and octant_name 
						for i in range(13,22):                                
							sheet.cell(row=j+1,column=i+1).value=count1[i-13]
							sheet.cell(row=j+1,column=i+1).border=make_border
						count1.pop(0)                                         
						rank_rank=rank_list(count1)                           
						list_rank1.append(find_rank1(rank_rank))                
						matrix_rank.append(rank_rank)                            
						
						# Writing the columns of rank, rank1 and octant_name 
						for i in range(8):                                              
							sheet.cell(row=j+1,column=23+i).value=matrix_rank[j-3][i]
							sheet.cell(row=j+1,column=23+i).border=make_border
							if(matrix_rank[j-3][i]==1):
								sheet.cell(row=j+1,column=23+i).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
						
						sheet.cell(row=j+1,column=31).value=list_rank1[j-3]
						sheet.cell(row=j+1,column=32).value=octant_name_id_mapping[str(list_rank1[j-3])]
						sheet.cell(row=j+1,column=31).border=make_border
						sheet.cell(row=j+1,column=32).border=make_border
						
						j=j+1
						oct_range_matrix.append(count1)
						count1=[0]*9                                
														
				list_rank1.pop(0)                                         
				j+=4
				sheet.cell(row=j,column=29).value='Octant ID'        
				sheet.cell(row=j,column=30).value='Octant Name'
				sheet.cell(row=j,column=31).value='Count of Rank 1 Mod Values'
				sheet.cell(row=j,column=29).border=make_border                 
				sheet.cell(row=j,column=30).border=make_border
				sheet.cell(row=j,column=31).border=make_border
				
				# Writing the table of count of rank1 mod values
				for i in range(8):                              
					sheet.cell(row=j+1+i,column=29).value=opp_map_dict[i]
					sheet.cell(row=j+1+i,column=30).value=octant_name_id_mapping[str(opp_map_dict[i])]
					sheet.cell(row=j+1+i,column=31).value=count_rank1(list_rank1,opp_map_dict[i])
					sheet.cell(row=j+1+i,column=29).border=make_border
					sheet.cell(row=j+1+i,column=30).border=make_border
					sheet.cell(row=j+1+i,column=31).border=make_border
			
			def oct_lsubs_count_with_range():
				
				r=['Count','Longest Subsequence Length','Count']

				# Writing header of table      
				for i in range(3):                                     
					sheet.cell(row=3,column=45+i).value=r[i] 
					sheet.cell(row=3,column=45+i).border=make_border    

				octs=[]
				for i in range(2,10,2):                                    
					sheet.cell(row=i+2,column=45).value=i//2
					octs.append(i//2)
					sheet.cell(row=i+3,column=45).value=-(i//2) 
					octs.append(-i//2)
					sheet.cell(row=i+2,column=45).border=make_border
					sheet.cell(row=i+3,column=45).border=make_border                                
				
			
				map_dict={}                                              
				for i in range(0,4):                                            
					map_dict[i+1]=2*i+1-1
					map_dict[-(i+1)]=2*(i+1)-1

				# List for storing number of long subseq		
				count1=[0]*8                                             
				long_len=[0]*8                                    
				back=octant[0]
				l=1                                                           
				n=len(octant)
				temp_max=[0]                                                  
				ranges= [[] for x in repeat(None, 8)]            

				# Loop for finding number and length of long subseq
				for i in range(1,n+1):               
					if(i==n):                                        
						if(long_len[map_dict[back]]<l):                        
							long_len[map_dict[back]]=l
							count1[map_dict[back]]=1

							temp_max.append(df['T'][i-1])            
							ranges[map_dict[back]].clear()         
							ranges[map_dict[back]].append(temp_max)           
						elif(long_len[map_dict[back]]==l):
							count1[map_dict[back]]+=1
							temp_max.append(df['T'][i-1])
							ranges[map_dict[back]].append(temp_max) 	               
					elif(back==octant[i]):                                      
						l+=1

					else:                                                      
						if(long_len[map_dict[back]]<l):
							long_len[map_dict[back]]=l
							count1[map_dict[back]]=1
							ranges[map_dict[back]].clear()               
							temp_max.append(df['T'][i-1])        
							ranges[map_dict[back]].append(temp_max)                 
						elif(long_len[map_dict[back]]==l):
							count1[map_dict[back]]+=1
							temp_max.append(df['T'][i-1])
							ranges[map_dict[back]].append(temp_max)             
						temp_max=[df['T'][i]]                     
						l=1
						back=octant[i]                            

				for i in range(2,10):                                    
					sheet.cell(row=i+2,column=46).value=long_len[i-2]
					sheet.cell(row=i+2,column=47).value=count1[i-2]
					sheet.cell(row=i+2,column=46).border=make_border
					sheet.cell(row=i+2,column=47).border=make_border
				k=2                                                        
				sheet.cell(row=k+1,column=49).value='Octant ###'                  
				sheet.cell(row=k+1,column=50).value='Longest Subsequence Length'
				sheet.cell(row=k+1,column=51).value='Count'
				sheet.cell(row=k+1,column=49).border=make_border
				sheet.cell(row=k+1,column=50).border=make_border
				sheet.cell(row=k+1,column=51).border=make_border
				
				k+=2
				for i in range(8):
					sheet.cell(row=k,column=49).value=octs[i]            
					sheet.cell(row=k,column=50).value=long_len[i]
					sheet.cell(row=k,column=51).value=count1[i]
					sheet.cell(row=k+1,column=49).value='Time'             
					sheet.cell(row=k+1,column=50).value='From'
					sheet.cell(row=k+1,column=51).value='To'
					sheet.cell(row=k,column=49).border=make_border      
					sheet.cell(row=k,column=50).border=make_border
					sheet.cell(row=k,column=51).border=make_border
					sheet.cell(row=k+1,column=49).border=make_border
					sheet.cell(row=k+1,column=50).border=make_border
					sheet.cell(row=k+1,column=51).border=make_border
					x=ranges[i]
					k+=2
					for j in x:
						# Writing ranges in worksheet
						sheet.cell(row=k,column=50).value=j[0]     
						sheet.cell(row=k,column=51).value=j[1]
						sheet.cell(row=k,column=49).border=make_border  
						sheet.cell(row=k,column=50).border=make_border
						sheet.cell(row=k,column=51).border=make_border
						k+=1
				
			def octant_transition_count(mod=5000):
				j=1
				n=len(octant)
				sheet.cell(row=j,column=35).value='Overall Transition Count' 
				sheet.cell(row=j+3,column=34).value='From'
				sheet.cell(row=j+1,column=36).value='To'
				j+=2
				
				oct_range_matrix = [ [0]*9 for i in range(9)]                    
				
				# Storing header row and header column in the matrix
				for i in range(0,4):                                        
					oct_range_matrix[0][2*i+1]=(i+1)
					oct_range_matrix[0][2*(i+1)]=-(i+1)
				for i in range(0,9):
					oct_range_matrix[i][0]=oct_range_matrix[0][i]
				oct_range_matrix[0][0]='Octant #'
				map_dict={}                                             
				for i in range(0,4):
					map_dict[i+1]=2*i+1
					map_dict[-(i+1)]=2*(i+1)

				##### Defining function to find row column
				def find_rc(x,y):                     
					comp_head=[map_dict[x],map_dict[y]]
					return comp_head
				
				#### Defining function to find max element
				def max_ele(comp_head):
					temp_max=comp_head.copy()
					temp_max.pop(0)
					large=0
					for i in temp_max:
						if(large<i):
							large=i
					return large

				back=octant[0]

				# Filling overall transition matrix                                              
				for i in range(1,n):                       
					comp_head=find_rc(back,octant[i])                        
					oct_range_matrix[comp_head[0]][comp_head[1]]+=1
					back=octant[i]

				CodeYelow = "00FFFF00"

				# Writing the overall transition matrix in worksheet
				for i in range(9):                                  
					temp_comp_head=oct_range_matrix[i]
					large=max_ele(temp_comp_head)
					for k in range(13,22):
						sheet.cell(row=j+i,column=k+22).value=oct_range_matrix[i][k-13]
						sheet.cell(row=j+i,column=k+22).border=make_border
						if(i>0 and oct_range_matrix[i][k-13]==large):
							sheet.cell(row=j+i,column=k+22).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
						if(i!=0 and k!=13):
							oct_range_matrix[i][k-13]=0

				# temp-> No. of mod transition tables	
				temp_max=n//mod+1                                              
				j+=1

				# One iteration for each mod transition table
				for t in range(temp_max):                                   
					j+=11
					name1=''

					# Writing Table name
					sheet.cell(row=j,column=35).value='Mod Transition Count'     
					sheet.cell(row=j+3,column=34).value='From'
					sheet.cell(row=j+1,column=36).value='To'
					name1=str(t*mod)+'-'
					if((t+1)*mod-1>n-1):
						name1+=str(n-1)
					else:
						name1+=str((t+1)*mod-1)   
					sheet.cell(row=j+1,column=35).value=name1
					j+=2

					############ Incrementing matrix cell corresponding to transition values
					for i in range(t*mod,min(n-1,(t+1)*mod)):                 
						comp_head=find_rc(octant[i],octant[i+1])
						oct_range_matrix[comp_head[0]][comp_head[1]]+=1

					########### Writing the transition mod matrix
					for i in range(0,9):                                     
						temp_comp_head=oct_range_matrix[i]
						if(i>0):
							large=max_ele(temp_comp_head)
						for k in range(13,22):
							sheet.cell(row=j+i,column=k+22).value=oct_range_matrix[i][k-13]
							sheet.cell(row=j+i,column=k+22).border=make_border
							if(i>0 and oct_range_matrix[i][k-13]==large):
								sheet.cell(row=j+i,column=k+22).fill=PatternFill(start_color=CodeYelow,end_color=CodeYelow,fill_type="solid")
							if(i!=0 and k!=13):

							
								oct_range_matrix[i][k-13]=0                           

			### running octant transition function
			octant_transition_count(mod)
			### Calling name octant range function
			name_oct_range(mod)
			### Calling long subsequence function
			oct_lsubs_count_with_range()
			file=file[:-5]
			#### Defining file output path
			path_out='output/'+file+' cm_vel_octant_analysis_mod_'+str(mod)+'.xlsx'
			wb.save(path_out)
			
			data_1 = BytesIO()
			###saving output file
			wb.save(data_1)

			#### Creating download button for output file
			
			filename = file +'cm_vel_octant_analysis_mod_'+str(mod)+'.xlsx'
			st.text(filename)
			st.download_button(label="Download File", file_name=filename   , data=data_1)
			
		for file in os.listdir(path):
			a = path+"\\"+file
			workingFile(a,file,mod)

	##### header of webapage 
	st.header("OCTANT ANALYSIS")

	###### Taking input file from user
	val = st.number_input(label="Enter Mod Value", min_value=1, step=1)
	file = st.file_uploader("Please choose a file", type=["xlsx"])
	if os.path.exists("output"):
		pass

	else:
		a = pathlib.Path(_file_).parent.parent.resolve()
		b = os.path.join(a,"proj2") 
		filename = os.path.join(b,"output") 
		os.mkdir(filename)

	########## Compute button to process output file
	if st.button("COMPUTE"):
		a = pathlib.Path(__file__).parent.parent.resolve()
		d = os.path.join(a,"proj2\output")
		for f in os.listdir(d):
			os.remove(os.path.join(d, f))
		octant_analysis_single(val,file)

	######### Computing output files for input of files' folder
	direc = st.text_input(label="If you want to upload more than 1 file \n\n Please give path of input files' directory")
	if st.button("COMPUTE_multiple"):
		
		octant_analysis_mutiple(val,direc)



###Code
proj_octant_gui()

#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
