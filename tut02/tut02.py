def octant_transition_count(mod=5000):
 from openpyxl import load_workbook
 import os
 import numpy as np
 os.system('cls')

 time = []  # new list for time
 v = []     # new list for V
 u = []     # new list for U
 w = []     # new list for W
 vd = []    # list for difference of Vavg - V
 ud = []    # list for difference of Uavg -U
 wd = []    # list for difference of Wavg - W
 oct = []   # list for storing octant

 lwb=load_workbook("input_octant_transition_identify.xlsx")
 sheet=lwb["Sheet1"] 
 n=sheet.max_row
 for r in range(2,n+1):
  time.append(sheet.cell(row=r,column=1).value)
  u.append(sheet.cell(row=r,column=2).value)
  v.append(sheet.cell(row=r,column=3).value)
  w.append(sheet.cell(row=r,column=4).value)

 v_avg=np.mean(v, dtype=np.float64) #  calculate average of v
 u_avg=np.mean(u, dtype=np.float64) #  calculate average of u
 w_avg=np.mean(w, dtype=np.float64) #  calculate average of w

 
 for r in range(2,n+1):
  ud.append(float(sheet.cell(row=r,column=2).value)-u_mean)  # push row[1] - u_avg in ud list
  vd.append(float(sheet.cell(row=r,column=3).value)-v_avg)  # push row[2] - v_avg in vd list
  wd.append(float(sheet.cell(row=r,column=4).value)-w_avg)	 # push row[3] - w_avg in wd list 

 for i in range(0,n-1):# find octants
  if ((ud[i]>=0) and (vd[i]>=0) and (wd[i]>=0 )):
   oct.append(1)      # adding the octant in oct list after checking the condition of octant 1 to calculate overall octant 1 value
  elif((ud[i]>=0) and (vd[i]>=0) and (wd[i]<0 )):
   oct.append(-1)     # adding the octant in oct list after checking the condition of octant -1 to calculate overall octant -1 value
  elif((ud[i]<0) and (vd[i]>=0) and (wd[i]>=0 )):
   oct.append(2)      # adding the octant in oct list after checking the condition of octant 2 to calculate overall octant 2 value
  elif((ud[i]<0) and (vd[i]>=0) and (wd[i]<0 )):
   oct.append(-2)     # adding the octant in oct list after checking the condition of octant -2 to calculate overall octant -2 value
  elif((ud[i]>=0) and (vd[i]<0) and (wd[i]>=0 )):
   oct.append(4)      # adding the octant in oct list after checking the condition of octant 4 to calculate overall octant 4 value
  elif((ud[i]>=0) and (vd[i]<0) and (wd[i]<0 )):
   oct.append(-4)    # adding the octant in oct list after checking the condition of octant -4 to calculate overall octant -4 value
  elif((ud[i]<0) and (vd[i]<0) and (wd[i]>=0 )):
   oct.append(3)     # adding the octant in oct list after checking the condition of octant 3 to calculate overall octant 3 value
  else:
   oct.append(-3)    # adding the octant in oct list after checking the condition of octant -3 to calculate overall octant =3 value
     
 count1=0         # calculating total count of octant 1
 count3=0         # calculating total count of octant 2
 count2=0         # calculating total count of octant -1
 count4=0         # calculating total count of octant -2
 count5=0         # calculating total count of octant 3
 count6=0         # calculating total count of octant -3
 count7=0         # calculating total count of octant 4
 count8=0         # calculating total count of octant -4	 

 for i in range(0,n-1):# for loop to calculate count of octants
  if (oct[i]==1):
   count1=count1+1
  elif(oct[i]==-1):
   count2=count2+1
  elif(oct[i]==2):
   count3=count3+1
  elif(oct[i]==-2):
   count4=count4+1
  elif(oct[i]==3):
   count5=count5+1
  elif(oct[i]==-3):
   count6=count6+1    
  elif(oct[i]==4):
   count7=count7+1
  else:
   count8=count8+1

  octant1=0
  octant2=0
  octant3=0
  octant4=0
  octant5=0
  octant6=0
  octant7=0
  octant8=0

 oct1=0 #making variables for using in function 
 oct2=0 #making variables for using in function
 oct3=0 #making variables for using in function
 oct4=0 #making variables for using in function
 oct5=0 #making variables for using in function
 oct6=0 #making variables for using in function
 oct7=0 #making variables for using in function
 oct8=0 #making variables for using in function

 m= int((n-2)/mod) +1

 for i in range(m):   # for loop to calculate individual count of octants in range of mod
  s=mod*i
  for j in range(mod):	
   if(j+s<n-1):	                                    
    if (oct[j+s]==1):
     oct1=oct1+1       #calculate count of octant 1 in range of particular mod
    elif(oct[j+s]==-1):
     oct2=oct2+1       #calculate count of octant -1 in range of particular mod
    elif(oct[j+s]==2):
     oct3=oct3+1       #calculate count of octant 2 in range of particular mod
    elif(oct[j+s]==-2):
     oct4=oct4+1       #calculate count of octant -2 in range of particular mod
    elif(oct[j+s]==3):
     oct5=oct5+1       #calculate count of octant 3 in range of particular mod
    elif(oct[j+s]==-3):
     oct6=oct6+1       #calculate count of octant -3 in range of particular mod
    elif(oct[j+s]==4):
     oct7=oct7+1       #calculate count of octant 4 in range of particular mod
    elif(oct[j+s]==-4):
     oct8=oct8+1       #calculate count of octant -4 in range of particular mod
  octant1.append(oct1)  # push values of count of octant 1 in particular range of mod
  octant2.append(oct2)  # push values of count of octant -1 in particular range of mod
  octant3.append(oct3)  # push values of count of octant 2 in particular range of mod
  octant4.append(oct4)  # push values of count of octant -2 in particular range of mod
  octant5.append(oct5)  # push values of count of octant 3 in particular range of mod
  octant6.append(oct6)  # push values of count of octant -3 in particular range of mod
  octant7.append(oct7)  # push values of count of octant 4 in particular range of mod
  octant8.append(oct8)  # push values of count of octant -4 in particular range of mod
  oct1=0
  oct2=0
  oct3=0
  oct4=0
  oct5=0
  oct6=0
  oct7=0
  oct8=0

 r=[]
 required=[]
 over=[[0]*8,[0]*8,[0]*8,[0]*8,[0]*8,[0]*8,[0]*8,[0]*8]
 for i in range(0,m) :
  p=i*mod
  x1=[0]*8 #transition from 1 to other octants
  x2=[0]*8 #transition from -1 to other octants
  x3=[0]*8 #transition from 2 to other octants
  x4=[0]*8 #transition from -2 to other octants
  x5=[0]*8 #transition from 3 to other octants
  x6=[0]*8 #transition from -3 to other octants
  x7=[0]*8 #transition from 4 to other octants
  x8=[0]*8 #transition from -4 to other octants
  for j in range(0,mod):
   if j+p<=n-3:
    if oct[j+p]==1 and oct[j+p+1]==1 : #transition count from 1 to 1
      x1[0]=x1[0]+1
      over[0][0]=over[0][0]+1
    elif oct[j+p]==1 and oct[j+p+1]==-1: #transition count from 1 to -1
      x1[1]=x1[1]+1    
      over[0][1]=over[0][1]+1
    elif oct[j+p]==1 and oct[j+p+1]==2: #transition count from 1 to 2
      x1[2]=x1[2]+1
      over[0][2]=over[0][2]+1
    elif oct[j+p]==1 and oct[j+p+1]==-2: #transition count from 1 to -2
      x1[3]=x1[3]+1
      over[0][3]=over[0][3]+1
    elif oct[j+p]==1 and oct[j+1+p]==3: #transition count from 1 to 3
      x1[4]=x1[4]+1 
      over[0][4]=over[0][4]+1
    elif oct[j+p]==1 and oct[j+p+1]==-3: #transition count from 1 to -3
      x1[5]=x1[5]+1
      over[0][5]=over[0][5]+1
    elif oct[j+p]==1 and oct[j+1+p]==4: #transition count from 1 to 4
      x1[6]=x1[6]+1
      over[0][6]=over[0][6]+1
    elif oct[j+p]==1 and oct[j+p+1]==-4: #transition count from 1 to -4
      x1[7]=x1[7]+1
      over[0][7]=over[0][7]+1
  
    elif oct[j+p]==-1 and oct[j+p+1]==1 : #transition count from -1 to 1
      x2[0]=x2[0]+1
      over[1][0]=over[1][0]+1
    elif oct[j+p]==-1 and oct[j+p+1]==-1: #transition count from -1 to -1
      x2[1]=x2[1]+1    
      over[1][1]=over[1][1]+1
    elif oct[j+p]==-1 and oct[j+p+1]==2: #transition count from -1 to 2
      x2[2]=x2[2]+1
      over[1][2]=over[1][2]+1
    elif oct[j+p]==-1 and oct[j+p+1]==-2: #transition count from -1 to -2
      x2[3]=x2[3]+1
      over[1][3]=over[1][3]+1
    elif oct[j+p]==-1 and oct[j+p+1]==3: #transition count from -1 to 3
      x2[4]=x2[4]+1
      over[1][4]=over[1][4]+1 
    elif oct[j+p]==-1 and oct[j+1+p]==-3: #transition count from -1 to -3
      x2[5]=x2[5]+1
      over[1][5]=over[1][5]+1
    elif oct[j+p]==-1 and oct[j+p+1]==4: #transition count from -1 to 4
      x2[6]=x2[6]+1
      over[1][6]=over[1][6]+1
    elif oct[j+p]==-1 and oct[j+1+p]==-4: #transition count from -1 to -4
      x2[7]=x2[7]+1
      over[1][7]=over[1][7]+1
    
    elif oct[j+p]==2 and oct[j+p+1]==1 : #transition count from 2 to 1
      x3[0]=x3[0]+1
      over[2][0]=over[2][0]+1
    elif oct[j+p]==2 and oct[j+p+1]==-1: #transition count from 2 to -1
      x3[1]=x3[1]+1
      over[2][1]=over[2][1]+1    
    elif oct[j+p]==2 and oct[j+p+1]==2: #transition count from 2 to 2
      x3[2]=x3[2]+1
      over[2][2]=over[2][2]+1
    elif oct[j+p]==2 and oct[j+p+1]==-2: #transition count from 2 to -2
      x3[3]=x3[3]+1
      over[2][3]=over[2][3]+1
    elif oct[j+p]==2 and oct[j+p+1]==3: #transition count from 2 to 3
      x3[4]=x3[4]+1
      over[2][4]=over[2][4]+1 
    elif oct[j+p]==2 and oct[j+1+p]==-3: #transition count from 2 to -3
      x3[5]=x3[5]+1
      over[2][5]=over[2][5]+1
    elif oct[j+p]==2 and oct[j+p+1]==4: #transition count from 2 to 4
      x3[6]=x3[6]+1
      over[2][6]=over[2][6]+1
    elif oct[j+p]==2 and oct[j+1+p]==-4: #transition count from 2 to -4
      x3[7]=x3[7]+1
      over[2][7]=over[2][7]+1

    elif oct[j+p]==-2 and oct[j+p+1]==1 : #transition count from -2 to 1
      x4[0]=x4[0]+1
      over[3][0]=over[3][0]+1
    elif oct[j+p]==-2 and oct[j+p+1]==-1: #transition count from -2 to -1
      x4[1]=x4[1]+1
      over[3][1]=over[3][1]+1    
    elif oct[j+p]==-2 and oct[j+p+1]==2: #transition count from -2 to 2
      x4[2]=x4[2]+1
      over[3][2]=over[3][2]+1
    elif oct[j+p]==-2 and oct[j+p+1]==-2: #transition count from -2 to -2
      x4[3]=x4[3]+1
      over[3][3]=over[3][3]+1
    elif oct[j+p]==-2 and oct[j+p+1]==3: #transition count from -2 to 3
      x4[4]=x4[4]+1 
      over[3][4]=over[3][4]+1
    elif oct[j+p]==-2 and oct[j+1+p]==-3: #transition count from -2 to -3
      x4[5]=x4[5]+1
      over[3][5]=over[3][5]+1
    elif oct[j+p]==-2 and oct[j+p+1]==4: #transition count from -2 to 4
      x4[6]=x4[6]+1
      over[3][6]=over[3][6]+1
    elif oct[j+p]==-2 and oct[j+1+p]==-4: #transition count from -2 to -4
      x4[7]=x4[7]+1
      over[3][7]=over[3][7]+1  
  
    elif oct[j+p]==3 and oct[j+p+1]==1 : #transition count from 3 to 1
      x5[0]=x5[0]+1
      over[4][0]=over[4][0]+1
    elif oct[j+p]==3 and oct[j+p+1]==-1: #transition count from 3 to -1
      x5[1]=x5[1]+1
      over[4][1]=over[4][1]+1    
    elif oct[j+p]==3 and oct[j+p+1]==2: #transition count from 3 to 2
      x5[2]=x5[2]+1
      over[4][2]=over[4][2]+1
    elif oct[j+p]==3 and oct[j+p+1]==-2: #transition count from 3 to -2
      x5[3]=x5[3]+1
      over[4][3]=over[4][3]+1
    elif oct[j+p]==3 and oct[j+p+1]==3: #transition count from 3 to 3
      x5[4]=x5[4]+1 
      over[4][4]=over[4][4]+1
    elif oct[j+p]==3 and oct[j+1+p]==-3: #transition count from 3 to -3
      x5[5]=x5[5]+1
      over[4][5]=over[4][5]+1
    elif oct[j+p]==3 and oct[j+p+1]==4: #transition count from 3 to 4
      x5[6]=x5[6]+1
      over[4][6]=over[4][6]+1
    elif oct[j+p]==3 and oct[j+1+p]==-4: #transition count from 3 to -4
      x5[7]=x5[7]+1
      over[4][7]=over[4][7]+1 
    
    elif oct[j+p]==-3 and oct[j+p+1]==1 : #transition count from -3 to 1
      x6[0]=x6[0]+1
      over[5][0]=over[5][0]+1
    elif oct[j+p]==-3 and oct[j+p+1]==-1: #transition count from -3 to -1
      x6[1]=x6[1]+1    
      over[5][1]=over[5][1]+1
    elif oct[j+p]==-3 and oct[j+p+1]==2: #transition count from -3 to 2
      x6[2]=x6[2]+1
      over[5][2]=over[5][2]+1
    elif oct[j+p]==-3 and oct[j+p+1]==-2: #transition count from -3 to -2
      x6[3]=x6[3]+1
      over[5][3]=over[5][3]+1
    elif oct[j+p]==-3 and oct[j+p+1]==3: #transition count from -3 to 3
      x6[4]=x6[4]+1
      over[5][4]=over[5][4]+1 
    elif oct[j+p]==-3 and oct[j+1+p]==-3: #transition count from -3 to -3
      x6[5]=x6[5]+1
      over[5][5]=over[5][5]+1
    elif oct[j+p]==-3 and oct[j+p+1]==4: #transition count from -3 to 4
      x6[6]=x6[6]+1
      over[5][6]=over[5][6]+1
    elif oct[j+p]==-3 and oct[j+1+p]==-4: #transition count from -3 to -4
      x6[7]=x6[7]+1
      over[5][7]=over[5][7]+1

    elif oct[j+p]==4 and oct[j+p+1]==1 : #transition count from 4 to 1
      x7[0]=x7[0]+1
      over[6][0]=over[6][0]+1
    elif oct[j+p]==4 and oct[j+p+1]==-1: #transition count from 4 to -1
      x7[1]=x7[1]+1    
      over[6][1]=over[6][1]+1
    elif oct[j+p]==4 and oct[j+p+1]==2: #transition count from 4 to 2
      x7[2]=x7[2]+1
      over[6][2]=over[6][2]+1
    elif oct[j+p]==4 and oct[j+p+1]==-2: #transition count from 4 to -2
      x7[3]=x7[3]+1
      over[6][3]=over[6][3]+1
    elif oct[j+p]==4 and oct[j+p+1]==3: #transition count from 4 to 3
      x7[4]=x7[4]+1
      over[6][4]=over[6][4]+1
    elif oct[j+p]==4 and oct[j+1+p]==-3: #transition count from 4 to -3
      x7[5]=x7[5]+1
      over[6][5]=over[6][5]+1
    elif oct[j+p]==4 and oct[j+p+1]==4: #transition count from 4 to 4
      x7[6]=x7[6]+1
      over[6][6]=over[6][6]+1
    elif oct[j+p]==4 and oct[j+1+p]==-4: #transition count from 4 to -4
      x7[7]=x7[7]+1 
      over[6][7]=over[6][7]+1


    elif oct[j+p]==-4 and oct[j+p+1]==1 : #transition count from -4 to 1
      x8[0]=x8[0]+1
      over[7][0]=over[7][0]+1
    elif oct[j+p]==-4 and oct[j+p+1]==-1: #transition count from -4 to -1
      x8[1]=x8[1]+1 
      over[7][1]=over[7][1]+1  
    elif oct[j+p]==-4 and oct[j+p+1]==2: #transition count from -4 to 2
      x8[2]=x8[2]+1
      over[7][2]=over[7][2]+1
    elif oct[j+p]==-4 and oct[j+p+1]==-2: #transition count from -4 to -2
      x8[3]=x8[3]+1
      over[7][3]=over[7][3]+1
    elif oct[j+p]==-4 and oct[j+p+1]==3: #transition count from -4 to 3
      x8[4]=x8[4]+1
      over[7][4]=over[7][4]+1 
    elif oct[j+p]==-4 and oct[j+1+p]==-3: #transition count from -4 to -3
      x8[5]=x8[5]+1
      over[7][5]=over[7][5]+1
    elif oct[j+p]==-4 and oct[j+p+1]==4: #transition count from -4 to 4
      x8[6]=x8[6]+1
      over[7][6]=over[7][6]+1
    elif oct[j+p]==-4 and oct[j+1+p]==-4: #transition count from -4 to -4
      x8[7]=x8[7]+1  
      over[7][7]=over[7][7]+1    
  if i==0 :
    r.append(over) #Adding all transitions in required
  required.append(x1)  
  required.append(x2) 
  required.append(x3)
  required.append(x4)
  required.append(x5)
  required.append(x6)
  required.append(x7)
  required.append(x8)
  c=required.copy()
  r.append(c)
  required.clear()
 h = ["+1","-1","+2","-2","+3","-3","+4","-4"]
 
 from openpyxl import Workbook 
 book=Workbook() #plot sheet of rows
 sheet= book.active    
 rows=[] 
 rows.append(["time",'U','V','W','Uavg','Vavg','Wavg',"U'=U-Uavg","V'=V-Vavg","W'=W-Wavg","Octant","","OctantID","+1","-1","+2","-2","+3","-3","+4","-4"])
 j=0
 a=0
 b=0
 for x in range(n-2):
  if(x==0):
   rows.append([time[x],u[x],v[x],w[x],u_mean,v_avg,w_avg,ud[x],vd[x],wd[x],oct[x],"","Overall count",count1,count2,count3,count4,count5,count6,count7,count8])
  elif(x==1):
   s="mod "+str(mod)		
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"User input",s,"","","","","","","",""])
  elif(x>=2 and x<2+m):
   if(x==1+m):# it will work ont=ly if x=1+m
    z=j*mod 	
    y=(j+1)*mod
    s=str(z)+"-"+str(n-2)		 
    rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"",s,octant1[x-2],octant2[x-2],octant3[x-2],octant4[x-2],octant5[x-2],octant6[x-2],octant7[x-2],octant8[x-2]])	 
    rows.append([time[x],u[x],v[x],w[x],u_mean,v_avg,w_avg,ud[x],vd[x],wd[x],oct[x],"","Verified",count1,count2,count3,count4,count5,count6,count7,count8])
    j=j+1
   else:
    z=j*mod	
    y=(j+1)*mod-1
    s=str(z)+"-"+str(y)		 
    rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"",s,octant1[x-2],octant2[x-2],octant3[x-2],octant4[x-2],octant5[x-2],octant6[x-2],octant7[x-2],octant8[x-2]])	 
    j=j+1
  elif (x-(2+m))%9==0 and x<9*(m+1)+2+m:
    rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","","","","","","","","",""])
    rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","","","","","","","","",""])
    if x ==2+m:
     rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","Overall transition Count","","","","","","","",""])
     rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","","To","","","","","","",""])
    else:
     rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","Mod transition Count","","","","","","","",""])
     rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"",str(((x-(2+m))//9-1)*mod)+"-"+str(np.minimum(((x-(2+m))//9)*mod,n-2)),"To","","","","","","",""])
    rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","Count","+1","-1","+2","-2","+3","-3","+4","-4"])
  elif x<9*(m+1)+(2+m):
   if (x-(2+m))%9==1:
    print("hi",b," ",a)
    rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"From",h[((x-(2+m))%9)-1],r[b][a][0],r[b][a][1],r[b][a][2],r[b][a][3],r[b][a][4],r[b][a][5],r[b][a][6],r[b][a][7]])	
   else:
    print("hi",b," ",a)
    rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"",h[((x-(2+m))%9)-1],r[b][a][0],r[b][a][1],r[b][a][2],r[b][a][3],r[b][a][4],r[b][a][5],r[b][a][6],r[b][a][7]])
   a=a+1
   if a==8:
    a=0
    b=b+1
  else:
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","","","","","","","",""])

 

 for i in rows:
  sheet.append(i)
 book.save("output_octant_transition_identify1.xlsx")
 
      
 
 


mod=3000
octant_transition_count(mod)