#Help https://youtu.be/H37f_x4wAC0
from datetime import datetime
start_time = datetime.now()

def octant_longest_subsequence_count_with_range():
###Code

 from openpyxl import load_workbook
 import os
 import numpy as np
 os.system('cls')

 time = [] #  new list for time
 v = []    #  new list for v
 u = []    #  new list for u
 w = []    #  new list for w
 vd = []    # list for difference of Vavg-v
 ud = []    # list for difference of Uavg-u
 wd = []    # list for difference of Wavg-w
 oct = []   # list for storing octant

 lwb=load_workbook("input_octant_longest_subsequence_with_range.xlsx")
 sheet=lwb["Sheet1"] 
 n=sheet.max_row
 for r in range(2,n+1):
  time.append(sheet.cell(row=r,column=1).value)
  u.append(sheet.cell(row=r,column=2).value)
  v.append(sheet.cell(row=r,column=3).value)
  w.append(sheet.cell(row=r,column=4).value)

 v_avg=np.mean(v, dtype=np.float64) # calculate  average of u
 u_avg=np.mean(u, dtype=np.float64) # calculate average of v
 w_avg=np.mean(w, dtype=np.float64) # calculate  average of w
 
 for r in range(2,n+1):
  ud.append(float(sheet.cell(row=r,column=2).value)-u_avg) # pushing  row[1]-Uavg in ud list
  vd.append(float(sheet.cell(row=r,column=3).value)-v_avg) # pushing row[2]-Vavg in vd list
  wd.append(float(sheet.cell(row=r,column=4).value)-w_avg)	# pushing  row[3]-Wavg in wd list 

 for i in range(0,n-1):# find octants
  if ((ud[i]>=0) and (vd[i]>=0) and (wd[i]>=0 )):
   oct.append(1)   # adding the octant in the oct list after checking the condition of octant 1 to calculate overall octant 1 value     
  elif((ud[i]>=0) and (vd[i]>=0) and (wd[i]<0 )):
   oct.append(-1)   # adding the octant in the oct list after checking the condition of octant -1 to calculate overall octant -1 value
  elif((ud[i]<0) and (vd[i]>=0) and (wd[i]>=0 )):
   oct.append(2)   # adding the octant in the oct list after checking the condition of octant 2 to calculate overall octant 2 value
  elif((ud[i]<0) and (vd[i]>=0) and (wd[i]<0 )):
   oct.append(-2)   # adding the octant in the oct list after checking the condition of octant -2 to calculate overall octant -2 value
  elif((ud[i]>=0) and (vd[i]<0) and (wd[i]>=0 )):
   oct.append(4)   # adding the octant in the oct list after checking the condition of octant 4 to calculate overall octant 4 value
  elif((ud[i]>=0) and (vd[i]<0) and (wd[i]<0 )):
   oct.append(-4)    # adding the octant in the oct list after checking the condition of octant -4 to calculate overall octant -4 value   
  elif((ud[i]<0) and (vd[i]<0) and (wd[i]>=0 )):
   oct.append(3)   # adding the octant in the oct list after checking the condition of octant 3 to calculate overall octant 3 value
  else:
   oct.append(-3)   # adding the octant in the oct list after checking the condition of octant -3 to calculate overall octant -3 value
 subs1=0 # defining a variable for storing longest susequece of octant 1
 subs2=0 #defining a variable for storing longest susequece of octant -1
 subs3=0 #defining a variable for storing longest susequece of octant 2
 subs4=0 #defining a variable for storing longest susequece of octant -2
 subs5=0 #defining a variable for storing longest susequece of octant 3
 subs6=0 #defining a variable for storing longest susequece of octant -3
 subs7=0 #defining a variable for storing longest susequece of octant 4
 subs8=0 #defining a variable for storing longest susequece of octant -4

 lsubs1=0 # defining a variable for storing longest subsequence length for octant 1
 lsubs2=0 # defining a variable for storing longest subsequence length for octant -1
 lsubs3=0 # defining a variable for storing longest subsequence length for octant 2
 lsubs4=0 # defining a variable for storing longest subsequence length for octant -2
 lsubs5=0 # defining a variable for storing longest subsequence length for octant 3
 lsubs6=0 # defining a variable for storing longest subsequence length for octant -3
 lsubs7=0 # defining a variable for storing longest subsequence length for octant 4
 lsubs8=0 # defining a variable for storing longest subsequence length for octant -4

 count1=0 # defining variable for count of octant 1
 count2=0 # defining variable for count of octant -1
 count3=0 # defining variable for count of octant 2
 count4=0 # defining variable for count of octant -2
 count5=0 # defining variable for count of octant 3
 count6=0 # defining variable for count of octant -3
 count7=0 # defining variable for count of octant 4
 count8=0 # defining variable for count of octant -4
 
 
 for r in range(0,n-1): # it will give count and longest subsequence length for octant 1
  if oct[r]==1 :
   subs1=subs1+1
   if r==n-2:
    if subs1>lsubs1: #checking if subsequence 1 is longest subsequence for octant 1
     lsubs1=subs1
     subs1=0
     count1=1        # if it is greater then upgrade the count1 to 1
    elif lsubs1>subs1:
     subs1=0        # if subs1 is not greater then assign it 0 value
    else:
     subs1=0
     count1=count1+1  # upgrade count1 by 1
  else:
    if subs1>lsubs1:
     lsubs1=subs1
     subs1=0
     count1=1
    elif lsubs1>subs1:
     subs1=0
    else:
     subs1=0
     count1=count1+1 # upgrade count1 by 1

 countcheck=0 # initiallize count check with 0
 time1=[] # This list will store the starting time for longest subsequence length for octant +1 
 for r in range(0,n-1):  
  if oct[r]==1:
   if countcheck==0: 
    ti1=time[r] 
   countcheck=countcheck+1
  else:
   countcheck=0 
  if countcheck==lsubs1: # Checking if length is equal to longest subsequence length
   time1.append(ti1); 
   countcheck=0

  
 
 for r in range(0,n-1): # it will give count and longest subsequence length for octant -1
  if oct[r]==-1 :
   subs2=subs2+1
   if r==n-2:
    if subs2>lsubs2:    #checking if subsequence 2 is longest subsequence for octant -1
     lsubs2=subs2
     subs2=0
     count2=1          # if it is greater then upgrade the count2 to 1
    elif lsubs2>subs2:
     subs2=0          # if subs2 is not greater then assign it 0 value
    else:
     subs2=0
     count2=count2+1  # upgrade count2 by 1
  else:
    if subs2>lsubs2:
     lsubs2=subs2
     subs2=0
     count2=1
    elif lsubs2>subs2:
     subs2=0
    else:
     subs2=0
     count2=count2+1  # upgrade count2 by 1

 countcheck=0 # initiallize count check with 0
 time2=[] # This list will store the starting time for longest subsequence length for octant -1
 for r in range(0,n-1):  
  if oct[r]==-1:
   if countcheck==0: 
    ti2=time[r] 
   countcheck=countcheck+1
  else:
   countcheck=0 
  if countcheck==lsubs2: # Checking if length is equal to longest subsequence length 
   time2.append(ti2); 
   countcheck=0

 
       

 for r in range(0,n-1): # This for loop will solve the count and longest subsequence length for octant 2
  if oct[r]==2 :
   subs3=subs3+1
   if r==n-2:
    if subs3>lsubs3:   #checking if subsequence 3 is longest subsequence for octant 2
     lsubs3=subs3
     subs3=0
     count3=1         # if it is greater then upgrade the count3 to 1
    elif lsubs3>subs3:
     subs3=0          # if subs3 is not greater then assign it 0 value
    else:
     subs3=0
     count3=count3+1  # upgrade count3 by 1
  else:
    if subs3>lsubs3:
     lsubs3=subs3
     subs3=0
     count3=1
    elif lsubs3>subs3:
     subs3=0
    else:
     subs3=0
     count3=count3+1  # upgrade count3 by 1 
 
 countcheck=0 # initiallize count check with 0
 time3=[]   # This list will store the starting time for longest subsequence length for octant 2
 for r in range(0,n-1):  
  if oct[r]==2:
   if countcheck==0: 
    ti3=time[r] 
   countcheck=countcheck+1
  else:
   countcheck=0 
  if countcheck==lsubs3: # Checking if length is equal to longest subsequence length 
   time3.append(ti3); 
   countcheck=0

 
 
 for r in range(0,n-1): # it will give count and longest subsequence length for octant -2
  if oct[r]==-2 :
   subs4=subs4+1
   if r==n-2:
    if subs4>lsubs4:    #checking if subsequence 4 is longest subsequence for octant -2
     lsubs4=subs4
     subs4=0
     count4=1          # if it is greater then upgrade the count4 to 1
    elif lsubs4>subs4:
     subs4=0           # if subs4 is not greater then assign it 0 value
    else:
     subs4=0
     count4=count4+1   # upgrade count4 by 1
  else:
    if subs4>lsubs4:
     lsubs4=subs4
     subs4=0
     count4=1 
    elif lsubs4>subs4:
     subs4=0
    else:
     subs4=0
     count4=count4+1    # upgrade count4 by 1

 countcheck=0 # initiallize with 0
 time4=[]  # This list will store the starting time for longest subsequence length for octant -2
 for r in range(0,n-1):  
  if oct[r]==-2:
   if countcheck==0: 
    ti4=time[r] 
   countcheck=countcheck+1
  else:
   countcheck=0 

  if countcheck==lsubs4: # Checking if length is equal to longest subsequence length 
   time4.append(ti4); 
   countcheck=0

  
     
 for r in range(0,n-1): # it will give count and longest subsequence length for octant 3
  if oct[r]==3 :
   subs5=subs5+1
   if r==n-2:
    if subs5>lsubs5:    #checking if subsequence 5 is longest subsequence for octant 3
     lsubs5=subs5
     subs5=0
     count5=1           # if it is greater then upgrade the count5 to 1
    elif lsubs5>subs5:
     subs5=0           # if subs5 is not greater then assign it 0 value
    else:
     subs5=0
     count5=count5+1   # upgrade count5 by 1
  else:
    if subs5>lsubs5:
     lsubs5=subs5
     subs5=0
     count5=1
    elif lsubs5>subs5:
     subs5=0
    else:
     subs5=0
     count5=count5+1  # upgrade count5 by 1
 
 countcheck=0 # initiallize count check with 0
 time5=[] # This list will store the starting time for longest subsequence length for octant 3
 for r in range(0,n-1):  
  if oct[r]==3:
   if countcheck==0: 
    ti5=time[r] 
   countcheck=countcheck+1
  else:
   countcheck=0 
  if countcheck==lsubs5: # Checking if length is equal to longest subsequence length
   time5.append(ti5); 
   countcheck=0

 
 
 for r in range(0,n-1): # it will give count and longest subsequence length for octant -3
  if oct[r]==-3 :
   subs6=subs6+1
   if r==n-2:
    if subs6>lsubs1:    #checking if subsequence 6 is longest subsequence for octant -3
     lsubs6=subs6
     subs6=0
     count6=1           # if it is greater then upgrade the count6 to 1
    elif lsubs6>subs6:
     subs6=0            # if subs6 is not greater then assign it 0 value
    else:
     subs6=0
     count6=count6+1    # upgrade count6 by 1
  else:
    if subs6>lsubs6:
     lsubs6=subs6
     subs6=0
     count6=1
    elif lsubs6>subs6:
     subs6=0
    else:
     subs6=0
     count6=count6+1  # upgrade count6 by 1
 
 countcheck=0 # initiallize count check with 0
 time6=[] # This list will store the starting time for longest subsequence length for octant -3
 for r in range(0,n-1):  
  if oct[r]==-3:
   if countcheck==0: 
    ti6=time[r] 
   countcheck=countcheck+1
  else:
   countcheck=0 
  if countcheck==lsubs6: # Checking if length is equal to longest subsequence length
   time6.append(ti6); 
   countcheck=0

 
 
 for r in range(0,n-1): # it will give count and longest subsequence length for octant 4
  if oct[r]==4 :
   subs7=subs7+1
   if r==n-2:
    if subs7>lsubs7:    #checking if subsequence 7 is longest subsequence for octant 4
     lsubs7=subs7
     subs7=0
     count7=1           # if it is greater then upgrade the count7 to 1
    elif lsubs7>subs7:
     subs7=0           # if subs7 is not greater then assign it 0 value
    else:
     subs7=0
     count7=count7+1   # upgrade count7 by 1
  else:
    if subs7>lsubs7:
     lsubs7=subs7
     subs7=0
     count7=1
    elif lsubs7>subs7:
     subs7=0
    else:
     subs7=0
     count7=count7+1   # upgrade count7 by 1

 countcheck=0 # initiallize count check with 0
 time7=[] # This list will store the starting time for longest subsequence length for octant 4
 for r in range(0,n-1):  
  if oct[r]==4:
   if countcheck==0: 
    ti7=time[r] 
   countcheck=countcheck+1
  else:
   countcheck=0 
  if countcheck==lsubs7: # Checking if length is equal to longest subsequence length
   time7.append(ti7); 
   countcheck=0

 
 
 for r in range(0,n-1): # it will give count and longest subsequence length for octant -4
  if oct[r]==-4 :
   subs8=subs8+1
   if r==n-2:
    if subs8>lsubs1:    #checking if subsequence 8 is longest subsequence for octant -4
     lsubs8=subs8
     subs8=0
     count8=1          # if it is greater then upgrade the count8 to 1
    elif lsubs8>subs8:
     subs8=0           # if subs8 is not greater then assign it 0 value
    else:
     subs8=0
     count8=count8+1   # upgrade count8 by 1
  else:
    if subs8>lsubs8:
     lsubs8=subs8
     subs8=0
     count8=1
    elif lsubs8>subs8:
     subs8=0
    else:
     subs8=0
     count8=count8+1  # upgrade count8 by 1
       
 countcheck=0 # initiallize count check with 0
 time8=[] # This list will store the starting time for longest subsequence length for octant -4
 for r in range(0,n-1):  
  if oct[r]==-4:
   if countcheck==0: 
    ti8=time[r] 
   countcheck=countcheck+1
  else:
   countcheck=0 
  if countcheck==lsubs8: # Checking if length is equal to longest subsequence length
   time8.append(ti8); 
   countcheck=0

 

 column1=[] # list for first column of required table
 column1.append("Octant")  
 column1.append("1")
 column1.append("Time")
 for i in range(0,count1): # give count1 no. of space in first column
  column1.append("")
 column1.append("-1")
 column1.append("Time")
 for i in range(0,count2): # give count2 no. of space in first column
  column1.append("")
 column1.append("2")
 column1.append("Time")
 for i in range(0,count3): # give count3 no. of space in first column
  column1.append("")
 column1.append("-2")
 column1.append("Time")
 for i in range(0,count4): # give count4 no. of space in first column
  column1.append("")
 column1.append("3")
 column1.append("Time")
 for i in range(0,count5): # give count5 no. of space in first column
  column1.append("")
 column1.append("-3")
 column1.append("Time")
 for i in range(0,count6): # give count6 no. of space in first column
  column1.append("")
 column1.append("4")
 column1.append("Time")
 for i in range(0,count7): # give count7 no. of space in first column
  column1.append("")
 column1.append("-4")
 column1.append("Time")
 for i in range(0,count8): # give count8 no. of space in first column
  column1.append("")
 v=len(column1)

 column2=[] # list for second column of required table
 column2.append("Longest Subsequence Length")
 column2.append(lsubs1)
 column2.append("From")
 for i in range(0,count1): # give starting time of longest subsequece length for octant +1 in 
  column2.append(time1[i])
 column2.append(lsubs2)
 column2.append("From")
 for i in range(0,count2): # give starting time of longest subsequece length for octant -1 in 
  column2.append(time2[i])
 column2.append(lsubs3)
 column2.append("From")
 for i in range(0,count3): # give starting time of longest subsequece length for octant +2 in
  column2.append(time3[i])
 column2.append(lsubs4)
 column2.append("From")
 for i in range(0,count4): # give starting time of longest subsequece length for octant -2 in
  column2.append(time4[i])
 column2.append(lsubs5)
 column2.append("From")
 for i in range(0,count5): # give starting time of longest subsequece length for octant +3 in
  column2.append(time5[i])
 column2.append(lsubs6)
 column2.append("From")
 for i in range(0,count6): # give starting time of longest subsequece length for octant -3 in
  column2.append(time6[i])
 column2.append(lsubs7)
 column2.append("From")
 for i in range(0,count7): # give starting time of longest subsequece length for octant +4 in
  column2.append(time7[i])
 column2.append(lsubs8)
 column2.append("From")
 for i in range(0,count8): # give starting time of longest subsequece length for octant -4 in
  column2.append(time8[i])

 column3=[] # list for third column of required table
 column3.append("Count")
 column3.append(count1)
 column3.append("To")
 for i in range(0,count1): # give ending time of longest subsequece length for octant +1 
  column3.append(time1[i]+ 0.01*(lsubs1-1))
 column3.append(count2)
 column3.append("To")
 for i in range(0,count2): # give ending time of longest subsequece length for octant -1 
  column3.append(time2[i]+ 0.01*(lsubs2-1))
 column3.append(count3)
 column3.append("To")
 for i in range(0,count3): # give ending time of longest subsequece length for octant +2 
  column3.append(time3[i]+ 0.01*(lsubs3-1))
 column3.append(count4)
 column3.append("To")
 for i in range(0,count4): # give ending time of longest subsequece length for octant -2 
  column3.append(time4[i]+ 0.01*(lsubs4-1))
 column3.append(count5)
 column3.append("To")
 for i in range(0,count5): # give ending time of longest subsequece length for octant +3 
  column3.append(time5[i]+ 0.01*(lsubs5-1))
 column3.append(count6)
 column3.append("To")
 for i in range(0,count6): # give ending time of longest subsequece length for octant -3 
  column3.append(time6[i]+ 0.01*(lsubs6-1))
 column3.append(count7)
 column3.append("To")
 for i in range(0,count7): # give ending time of longest subsequece length for octant +4 
  column3.append(time7[i]+ 0.01*(lsubs7-1))
 column3.append(count8)
 column3.append("To")
 for i in range(0,count8): # give ending time of longest subsequece length for octant +4 
  column3.append(time8[i]+ 0.01*(lsubs8-1))

#plotting the workbook
 from openpyxl import Workbook 
 book=Workbook()
 sheet= book.active    
 rows=[] 
 rows.append(["Time",'U','V','W','Uavg','Vavg','Wavg',"U'=U-Uavg","V'=V-Vavg","W'=W-Wavg","Octant"])
 for x in range(n-2): # appending the data in xlsx file 
  if(x==0):
  
   rows.append([time[x],u[x],v[x],w[x],u_avg,v_avg,w_avg,ud[x],vd[x],wd[x],oct[x],"","Octant","Longest Susequence Length","Count","","Octant","Longest Susequence Length","Count"])
  elif x==1: # appending the data in xlsx file for octant 1  and time ranges 
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","1",lsubs1,count1,"",column1[x],column2[x],column3[x]])
  elif x==2: # appending the data in xlsx file for octant -1 and time ranges 
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","-1",lsubs2,count2,"",column1[x],column2[x],column3[x]])
  elif x==3:  # appending the data in xlsx file for octant 2 and time ranges
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","2",lsubs3,count3,"",column1[x],column2[x],column3[x]])
  elif x==4:  # appending the data in xlsx file for octant -2 and time ranges
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","-2",lsubs4,count4,"",column1[x],column2[x],column3[x]]) 
  elif x==5: # appending the data in xlsx file for octant 3 and time ranges 
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","3",lsubs5,count5,"",column1[x],column2[x],column3[x]])
  elif x==6: # appending the data in xlsx file for octant -3 and time ranges 
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","-3",lsubs6,count6,"",column1[x],column2[x],column3[x]])
  elif x==7: # appending the data in xlsx file for octant 4 and time ranges 
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","4",lsubs7,count7,"",column1[x],column2[x],column3[x]])
  elif x==8: # appending the data in xlsx file for octant -4 and time ranges 
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","-4",lsubs8,count8,"",column1[x],column2[x],column3[x]])
  elif x<v:
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","","","","",column1[x],column2[x],column3[x]])
  else: # appending the remaining data in xlsx file 
   rows.append([time[x],u[x],v[x], w[x],"","","",ud[x],vd[x],wd[x],oct[x]])

 for i in rows:
  sheet.append(i)
 book.save("output_octant_longest_subsequence_with_range.xlsx")
 

from platform import python_version
ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


octant_longest_subsequence_count_with_range()
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))