#Help https://youtu.be/H37f_x4wAC0
def octant_longest_subsequence_count():
 from openpyxl import load_workbook
 import os
 import numpy as np
 os.system('cls')

 time = [] #  new list for time
 v = []    #  new list for v
 u = []    #  new list for u
 w = []    #  new list for w
 vd = []    # list for difference of Vavg-v
 ud = []    # list for difference of Uavg-u
 wd = []    # list for difference of Wavg-w
 oct = []   # list for storing octant

 lwb=load_workbook("input_octant_longest_subsequence.xlsx")
 sheet=lwb["Sheet1"] 
 n=sheet.max_row
 for r in range(2,n+1):
  time.append(sheet.cell(row=r,column=1).value)
  u.append(sheet.cell(row=r,column=2).value)
  v.append(sheet.cell(row=r,column=3).value)
  w.append(sheet.cell(row=r,column=4).value)

 v_avg=np.mean(v, dtype=np.float64) # calculate  average of u
 u_avg=np.mean(u, dtype=np.float64) # calculate average of v
 w_avg=np.mean(w, dtype=np.float64) # calculate  average of w
 
 for r in range(2,n+1):
  ud.append(float(sheet.cell(row=r,column=2).value)-u_avg) # pushing  row[1]-Uavg in ud list
  vd.append(float(sheet.cell(row=r,column=3).value)-v_avg) # pushing row[2]-Vavg in vd list
  wd.append(float(sheet.cell(row=r,column=4).value)-w_avg)	# pushing  row[3]-Wavg in wd list 

 for i in range(0,n-1):# find octants
  if ((ud[i]>=0) and (vd[i]>=0) and (wd[i]>=0 )):
   oct.append(1)   # adding the octant in the oct list after checking the condition of octant 1 to calculate overall octant 1 value     
  elif((ud[i]>=0) and (vd[i]>=0) and (wd[i]<0 )):
   oct.append(-1)   # adding the octant in the oct list after checking the condition of octant -1 to calculate overall octant -1 value
  elif((ud[i]<0) and (vd[i]>=0) and (wd[i]>=0 )):
   oct.append(2)   # adding the octant in the oct list after checking the condition of octant 2 to calculate overall octant 2 value
  elif((ud[i]<0) and (vd[i]>=0) and (wd[i]<0 )):
   oct.append(-2)   # adding the octant in the oct list after checking the condition of octant -2 to calculate overall octant -2 value
  elif((ud[i]>=0) and (vd[i]<0) and (wd[i]>=0 )):
   oct.append(4)   # adding the octant in the oct list after checking the condition of octant 4 to calculate overall octant 4 value
  elif((ud[i]>=0) and (vd[i]<0) and (wd[i]<0 )):
   oct.append(-4)    # adding the octant in the oct list after checking the condition of octant -4 to calculate overall octant -4 value   
  elif((ud[i]<0) and (vd[i]<0) and (wd[i]>=0 )):
   oct.append(3)   # adding the octant in the oct list after checking the condition of octant 3 to calculate overall octant 3 value
  else:
   oct.append(-3)   # adding the octant in the oct list after checking the condition of octant -3 to calculate overall octant -3 value
 subs1=0 # defining a variable for storing longest susequece of octant 1
 subs2=0 #defining a variable for storing longest susequece of octant -1
 subs3=0 #defining a variable for storing longest susequece of octant 2
 subs4=0 #defining a variable for storing longest susequece of octant -2
 subs5=0 #defining a variable for storing longest susequece of octant 3
 subs6=0 #defining a variable for storing longest susequece of octant -3
 subs7=0 #defining a variable for storing longest susequece of octant 4
 subs8=0 #defining a variable for storing longest susequece of octant -4

 lsubs1=0 # defining a variable for storing longest subsequence length for octant 1
 lsubs2=0 # defining a variable for storing longest subsequence length for octant -1
 lsubs3=0 # defining a variable for storing longest subsequence length for octant 2
 lsubs4=0 # defining a variable for storing longest subsequence length for octant -2
 lsubs5=0 # defining a variable for storing longest subsequence length for octant 3
 lsubs6=0 # defining a variable for storing longest subsequence length for octant -3
 lsubs7=0 # defining a variable for storing longest subsequence length for octant 4
 lsubs8=0 # defining a variable for storing longest subsequence length for octant -4

 count1=0 # defining variable for count of octant 1
 count2=0 # defining variable for count of octant -1
 count3=0 # defining variable for count of octant 2
 count4=0 # defining variable for count of octant -2
 count5=0 # defining variable for count of octant 3
 count6=0 # defining variable for count of octant -3
 count7=0 # defining variable for count of octant 4
 count8=0 # defining variable for count of octant -4
 
 
 for r in range(0,n-1): # it will give count and longest subsequence length for octant 1
  if oct[r]==1 :
   subs1=subs1+1
   if r==n-2:
    if subs1>lsubs1: #checking if subsequence 1 is longest subsequence for octant 1
     lsubs1=subs1
     subs1=0
     count1=1        # if it is greater then upgrade the count1 to 1
    elif lsubs1>subs1:
     subs1=0        # if subs1 is not greater then assign it 0 value
    else:
     subs1=0
     count1=count1+1  # upgrade count1 by 1
  else:
    if subs1>lsubs1:
     lsubs1=subs1
     subs1=0
     count1=1
    elif lsubs1>subs1:
     subs1=0
    else:
     subs1=0
     count1=count1+1 # upgrade count1 by 1

 for r in range(0,n-1): # it will give count and longest subsequence length for octant -1
  if oct[r]==-1 :
   subs2=subs2+1
   if r==n-2:
    if subs2>lsubs2:    #checking if subsequence 2 is longest subsequence for octant -1
     lsubs2=subs2
     subs2=0
     count2=1          # if it is greater then upgrade the count2 to 1
    elif lsubs2>subs2:
     subs2=0          # if subs2 is not greater then assign it 0 value
    else:
     subs2=0
     count2=count2+1  # upgrade count2 by 1
  else:
    if subs2>lsubs2:
     lsubs2=subs2
     subs2=0
     count2=1
    elif lsubs2>subs2:
     subs2=0
    else:
     subs2=0
     count2=count2+1  # upgrade count2 by 1
 for r in range(0,n-1): # This for loop will solve the count and longest subsequence length for octant 2
  if oct[r]==2 :
   subs3=subs3+1
   if r==n-2:
    if subs3>lsubs3:   #checking if subsequence 3 is longest subsequence for octant 2
     lsubs3=subs3
     subs3=0
     count3=1         # if it is greater then upgrade the count3 to 1
    elif lsubs3>subs3:
     subs3=0          # if subs3 is not greater then assign it 0 value
    else:
     subs3=0
     count3=count3+1  # upgrade count3 by 1
  else:
    if subs3>lsubs3:
     lsubs3=subs3
     subs3=0
     count3=1
    elif lsubs3>subs3:
     subs3=0
    else:
     subs3=0
     count3=count3+1  # upgrade count3 by 1

 for r in range(0,n-1): # it will give count and longest subsequence length for octant -2
  if oct[r]==-2 :
   subs4=subs4+1
   if r==n-2:
    if subs4>lsubs4:    #checking if subsequence 4 is longest subsequence for octant -2
     lsubs4=subs4
     subs4=0
     count4=1          # if it is greater then upgrade the count4 to 1
    elif lsubs4>subs4:
     subs4=0           # if subs4 is not greater then assign it 0 value
    else:
     subs4=0
     count4=count4+1   # upgrade count4 by 1
  else:
    if subs4>lsubs4:
     lsubs4=subs4
     subs4=0
     count4=1 
    elif lsubs4>subs4:
     subs4=0
    else:
     subs4=0
     count4=count4+1    # upgrade count4 by 1
     
 for r in range(0,n-1): # it will give count and longest subsequence length for octant 3
  if oct[r]==3 :
   subs5=subs5+1
   if r==n-2:
    if subs5>lsubs5:    #checking if subsequence 5 is longest subsequence for octant 3
     lsubs5=subs5
     subs5=0
     count5=1           # if it is greater then upgrade the count5 to 1
    elif lsubs5>subs5:
     subs5=0           # if subs5 is not greater then assign it 0 value
    else:
     subs5=0
     count5=count5+1   # upgrade count5 by 1
  else:
    if subs5>lsubs5:
     lsubs5=subs5
     subs5=0
     count5=1
    elif lsubs5>subs5:
     subs5=0
    else:
     subs5=0
     count5=count5+1  # upgrade count5 by 1

 for r in range(0,n-1): # it will give count and longest subsequence length for octant -3
  if oct[r]==-3 :
   subs6=subs6+1
   if r==n-2:
    if subs6>lsubs1:    #checking if subsequence 6 is longest subsequence for octant -3
     lsubs6=subs6
     subs6=0
     count6=1           # if it is greater then upgrade the count6 to 1
    elif lsubs6>subs6:
     subs6=0            # if subs6 is not greater then assign it 0 value
    else:
     subs6=0
     count6=count6+1    # upgrade count6 by 1
  else:
    if subs6>lsubs6:
     lsubs6=subs6
     subs6=0
     count6=1
    elif lsubs6>subs6:
     subs6=0
    else:
     subs6=0
     count6=count6+1  # upgrade count6 by 1
 for r in range(0,n-1): # it will give count and longest subsequence length for octant 4
  if oct[r]==4 :
   subs7=subs7+1
   if r==n-2:
    if subs7>lsubs7:    #checking if subsequence 7 is longest subsequence for octant 4
     lsubs7=subs7
     subs7=0
     count7=1           # if it is greater then upgrade the count7 to 1
    elif lsubs7>subs7:
     subs7=0           # if subs7 is not greater then assign it 0 value
    else:
     subs7=0
     count7=count7+1   # upgrade count7 by 1
  else:
    if subs7>lsubs7:
     lsubs7=subs7
     subs7=0
     count7=1
    elif lsubs7>subs7:
     subs7=0
    else:
     subs7=0
     count7=count7+1   # upgrade count7 by 1

 for r in range(0,n-1): # it will give count and longest subsequence length for octant -4
  if oct[r]==-4 :
   subs8=subs8+1
   if r==n-2:
    if subs8>lsubs1:    #checking if subsequence 8 is longest subsequence for octant -4
     lsubs8=subs8
     subs8=0
     count8=1          # if it is greater then upgrade the count8 to 1
    elif lsubs8>subs8:
     subs8=0           # if subs8 is not greater then assign it 0 value
    else:
     subs8=0
     count8=count8+1   # upgrade count8 by 1
  else:
    if subs8>lsubs8:
     lsubs8=subs8
     subs8=0
     count8=1
    elif lsubs8>subs8:
     subs8=0
    else:
     subs8=0
     count8=count8+1  # upgrade count8 by 1
     
#plotting the workbook for output
 from openpyxl import Workbook  
 book=Workbook()    
 sheet= book.active    
 rows=[] 
 rows.append(["Time",'U','V','W','Uavg','Vavg','Wavg',"U'=U-Uavg","V'=V-Vavg","W'=W-Wavg","Octant"])
 for x in range(n-2): # appending the data in xlsx file 
  if(x==0):
   rows.append([time[x],u[x],v[x],w[x],u_avg,v_avg,w_avg,ud[x],vd[x],wd[x],oct[x],"","Octant","Longest Susequence Length","Count"])
  elif x==1: # appending the data in xlsx file for octant 1  
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","1",lsubs1,count1])
  elif x==2: # appending the data in xlsx file for octant -1 
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","-1",lsubs2,count2])
  elif x==3:  # appending the data in xlsx file for octant 2
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","2",lsubs3,count3])
  elif x==4:  # appending the data in xlsx file for octant -2
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","-2",lsubs4,count4]) 
  elif x==5: # appending the data in xlsx file for octant 3 
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","3",lsubs5,count5])
  elif x==6: # appending the data in xlsx file for octant -3 
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","-3",lsubs6,count6])
  elif x==7: # appending the data in xlsx file for octant 4 
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","4",lsubs7,count7])
  elif x==8: # appending the data in xlsx file for octant -4 
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x],"","-4",lsubs8,count8])
  else: # appending the remaining data in xlsx file 
   rows.append([time[x],u[x],v[x],w[x],"","","",ud[x],vd[x],wd[x],oct[x]])

 for i in rows:
  sheet.append(i)
 book.save("output_octant_longest_subsequence.xlsx")
       
from platform import python_version
ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


octant_longest_subsequence_count()
